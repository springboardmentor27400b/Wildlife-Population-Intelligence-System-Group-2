from fastapi import APIRouter, Depends, HTTPException, status, Query
from app.models.audit_log import AuditLog
from app.models.user import User
from app.api.auth import get_current_user
from typing import List, Optional
from datetime import datetime
from fastapi.responses import Response
import io
import openpyxl
from fpdf import FPDF
from beanie.operators import RegEx, GTE, LTE

router = APIRouter()

async def check_admin(current_user: User = Depends(get_current_user)):
    if current_user.role.lower() != "admin":
        raise HTTPException(status_code=403, detail="Not authorized. Admins only.")
    return current_user

def build_query(
    search: Optional[str] = None,
    module: Optional[str] = None,
    action: Optional[str] = None,
    status: Optional[str] = None,
    severity: Optional[str] = None,
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
    user: Optional[str] = None
):
    conditions = []
    
    if search:
        search_regex = {"$regex": search, "$options": "i"}
        conditions.append({
            "$or": [
                {"user_name": search_regex},
                {"description": search_regex},
                {"action": search_regex},
                {"module": search_regex}
            ]
        })
        
    if user:
        user_regex = {"$regex": user, "$options": "i"}
        conditions.append({
            "$or": [
                {"user_name": user_regex},
                {"user_id": user}
            ]
        })
        
    if module:
        conditions.append({"module": module})
    if action:
        conditions.append({"action": action})
    if status:
        conditions.append({"status": status})
    if severity:
        conditions.append({"severity": severity})
        
    if start_date or end_date:
        ts_cond = {}
        if start_date:
            ts_cond["$gte"] = start_date
        if end_date:
            ts_cond["$lte"] = end_date
        conditions.append({"timestamp": ts_cond})
        
    if not conditions:
        return {}
    if len(conditions) == 1:
        return conditions[0]
    return {"$and": conditions}

@router.get("/")
async def get_audit_logs(
    current_user: User = Depends(check_admin),
    page: int = Query(1, ge=1),
    limit: int = Query(20, ge=1, le=100),
    search: Optional[str] = None,
    module: Optional[str] = None,
    action: Optional[str] = None,
    status: Optional[str] = None,
    severity: Optional[str] = None,
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
    user: Optional[str] = None
):
    query = build_query(
        search=search,
        module=module,
        action=action,
        status=status,
        severity=severity,
        start_date=start_date,
        end_date=end_date,
        user=user
    )
    
    total = await AuditLog.find(query).count()
    skip = (page - 1) * limit
    
    logs = await AuditLog.find(query).sort("-timestamp").skip(skip).limit(limit).to_list()
    
    return {
        "total": total,
        "page": page,
        "limit": limit,
        "logs": logs
    }

@router.get("/export/excel")
async def export_audit_logs_excel(
    current_user: User = Depends(check_admin),
    search: Optional[str] = None,
    module: Optional[str] = None,
    action: Optional[str] = None,
    status: Optional[str] = None,
    severity: Optional[str] = None,
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
    user: Optional[str] = None
):
    query = build_query(
        search=search,
        module=module,
        action=action,
        status=status,
        severity=severity,
        start_date=start_date,
        end_date=end_date,
        user=user
    )
    
    logs = await AuditLog.find(query).sort("-timestamp").limit(1000).to_list()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Audit Logs"
    
    headers = ["Timestamp", "User", "Role", "Module", "Action", "Description", "IP Address", "Severity", "Status"]
    ws.append(headers)
    
    for log in logs:
        ws.append([
            log.timestamp.strftime("%Y-%m-%d %H:%M:%S") if log.timestamp else "",
            log.user_name,
            log.user_role,
            log.module,
            log.action,
            log.description,
            log.ip_address,
            log.severity,
            log.status
        ])
        
    # Auto filters
    ws.auto_filter.ref = f"A1:I{len(logs) + 1}"
    
    # Frozen header (freezes row 1)
    ws.freeze_panes = "A2"
    
    # Auto-sized columns
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 10)
        
    # Separate worksheet for summary statistics
    stats_ws = wb.create_sheet(title="Summary Statistics")
    
    # Calculate statistics
    total_logs = len(logs)
    
    severity_counts = {"INFO": 0, "WARNING": 0, "ERROR": 0, "SUCCESS": 0}
    status_counts = {"Success": 0, "Failed": 0}
    module_counts = {}
    
    for log in logs:
        sev = log.severity.upper() if log.severity else "INFO"
        severity_counts[sev] = severity_counts.get(sev, 0) + 1
        
        stat = log.status.capitalize() if log.status else "Success"
        status_counts[stat] = status_counts.get(stat, 0) + 1
        
        mod = log.module or "System"
        module_counts[mod] = module_counts.get(mod, 0) + 1
        
    stats_ws.append(["Summary Statistics", "Value"])
    stats_ws.append(["Total Log Entries", total_logs])
    stats_ws.append([])
    
    stats_ws.append(["Severity Breakdown", "Count"])
    for sev, count in severity_counts.items():
        stats_ws.append([sev, count])
    stats_ws.append([])
    
    stats_ws.append(["Status Breakdown", "Count"])
    for stat, count in status_counts.items():
        stats_ws.append([stat, count])
    stats_ws.append([])
    
    stats_ws.append(["Module Breakdown", "Count"])
    for mod, count in sorted(module_counts.items(), key=lambda x: x[1], reverse=True):
        stats_ws.append([mod, count])
        
    # Format Summary Statistics columns
    for col in stats_ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        stats_ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return Response(
        content=output.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=audit_logs.xlsx"}
    )

class AuditPDF(FPDF):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.alias_nb_pages()
        
    def footer(self):
        self.set_y(-15)
        self.set_font("helvetica", "I", 8)
        self.set_text_color(128, 128, 128)
        self.cell(0, 10, f"Page {self.page_no()}/{{nb}}", align="C")

@router.get("/export/pdf")
async def export_audit_logs_pdf(
    current_user: User = Depends(check_admin),
    search: Optional[str] = None,
    module: Optional[str] = None,
    action: Optional[str] = None,
    status: Optional[str] = None,
    severity: Optional[str] = None,
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
    user: Optional[str] = None
):
    query = build_query(
        search=search,
        module=module,
        action=action,
        status=status,
        severity=severity,
        start_date=start_date,
        end_date=end_date,
        user=user
    )
    
    logs = await AuditLog.find(query).sort("-timestamp").limit(1000).to_list()
    
    pdf = AuditPDF(orientation='L')
    pdf.add_page()
    
    # Title
    pdf.set_font("helvetica", "B", 16)
    pdf.set_text_color(31, 41, 55) # charcoal
    pdf.cell(0, 10, "WPIS - Audit Logs Report", ln=True, align="C")
    
    # Subtitle / Date
    pdf.set_font("helvetica", size=10)
    pdf.set_text_color(107, 114, 128) # grey
    pdf.cell(0, 8, f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", ln=True, align="C")
    
    # Applied Filters Section
    applied_filters = []
    if search: applied_filters.append(f"Search: '{search}'")
    if user: applied_filters.append(f"User: '{user}'")
    if module: applied_filters.append(f"Module: {module}")
    if action: applied_filters.append(f"Action: {action}")
    if severity: applied_filters.append(f"Severity: {severity}")
    if status: applied_filters.append(f"Status: {status}")
    if start_date: applied_filters.append(f"From: {start_date.strftime('%Y-%m-%d')}")
    if end_date: applied_filters.append(f"To: {end_date.strftime('%Y-%m-%d')}")
    
    filter_text = "Applied Filters: " + (", ".join(applied_filters) if applied_filters else "None")
    pdf.cell(0, 8, filter_text, ln=True, align="C")
    pdf.ln(5)
    
    # Table Header
    pdf.set_font("helvetica", "B", 10)
    pdf.set_fill_color(34, 197, 94) # Green (#22c55e)
    pdf.set_text_color(255, 255, 255) # White
    
    col_widths = [35, 35, 30, 35, 92, 25, 25]
    headers = ["Timestamp", "User", "Module", "Action", "Description", "Severity", "Status"]
    
    for i, header in enumerate(headers):
        pdf.cell(col_widths[i], 10, header, border=1, fill=True, align="L")
    pdf.ln()
    
    # Table Rows
    pdf.set_font("helvetica", size=9)
    pdf.set_text_color(31, 41, 55) # charcoal
    
    fill = False
    for log in logs:
        pdf.set_fill_color(249, 250, 251) # Alternate row background color
        
        timestamp_str = log.timestamp.strftime("%Y-%m-%d %H:%M:%S") if log.timestamp else ""
        user_name_str = log.user_name or "System/Anonymous"
        module_str = log.module or ""
        action_str = log.action or ""
        description_str = log.description or ""
        severity_str = log.severity or "INFO"
        status_str = log.status or "Success"
        
        # Truncate strings to prevent overlapping/wrapping issues
        user_name_str = user_name_str[:18] + ".." if len(user_name_str) > 20 else user_name_str
        module_str = module_str[:13] + ".." if len(module_str) > 15 else module_str
        action_str = action_str[:18] + ".." if len(action_str) > 20 else action_str
        description_str = description_str[:48] + ".." if len(description_str) > 50 else description_str
        
        pdf.cell(col_widths[0], 8, timestamp_str, border=1, fill=fill)
        pdf.cell(col_widths[1], 8, user_name_str, border=1, fill=fill)
        pdf.cell(col_widths[2], 8, module_str, border=1, fill=fill)
        pdf.cell(col_widths[3], 8, action_str, border=1, fill=fill)
        pdf.cell(col_widths[4], 8, description_str, border=1, fill=fill)
        pdf.cell(col_widths[5], 8, severity_str, border=1, fill=fill)
        pdf.cell(col_widths[6], 8, status_str, border=1, fill=fill)
        pdf.ln()
        fill = not fill
        
    pdf_bytes = pdf.output()
    return Response(
        content=bytes(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=audit_logs.pdf"}
    )

@router.get("/{id}")
async def get_audit_log(
    id: str,
    current_user: User = Depends(check_admin)
):
    from beanie import PydanticObjectId
    from bson.errors import InvalidId
    
    try:
        obj_id = PydanticObjectId(id)
    except InvalidId:
        raise HTTPException(status_code=400, detail="Invalid log ID format")
        
    log = await AuditLog.get(obj_id)
    if not log:
        raise HTTPException(status_code=404, detail="Log not found")
    return log
