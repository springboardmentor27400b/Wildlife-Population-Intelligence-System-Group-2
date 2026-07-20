import io
from fastapi import APIRouter, Depends, Query
from fastapi.responses import Response
from typing import Optional
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from fpdf import FPDF

from app.models.observation import ObservationRecord
from app.models.site import MonitoringSite
from app.models.device import SensorDevice
from app.models.upload import FieldUpload
from app.models.user import User
from app.models.notification import Notification
from app.models.prediction import PredictionRecord
from app.api.auth import get_current_user

router = APIRouter()


def build_filter_query(
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
    species: Optional[str] = None,
    monitoring_site_id: Optional[str] = None,
    verification_status: Optional[str] = None,
    observer_id: Optional[str] = None,
    search: Optional[str] = None,
) -> dict:
    query = {}
    if start_date or end_date:
        query["observed_at"] = {}
        if start_date:
            query["observed_at"]["$gte"] = start_date
        if end_date:
            query["observed_at"]["$lte"] = end_date
    if species:
        query["species_name"] = {"$regex": species, "$options": "i"}
    if monitoring_site_id:
        query["monitoring_site_id"] = monitoring_site_id
    if verification_status:
        query["verification_status"] = verification_status
    if observer_id:
        query["observer_id"] = observer_id
    if search:
        query["$or"] = [
            {"species_name": {"$regex": search, "$options": "i"}},
            {"observer_name": {"$regex": search, "$options": "i"}},
            {"monitoring_site_name": {"$regex": search, "$options": "i"}},
        ]
    return query


@router.get("/summary")
async def get_reports_summary(
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
    species: Optional[str] = None,
    monitoring_site_id: Optional[str] = None,
    verification_status: Optional[str] = None,
    observer_id: Optional[str] = None,
    search: Optional[str] = None,
    current_user: User = Depends(get_current_user),
):
    query = build_filter_query(
        start_date, end_date, species, monitoring_site_id, verification_status, observer_id, search
    )

    total_obs = await ObservationRecord.find(query).count()

    verified_query = {**query, "verification_status": "Verified"}
    verified_obs = await ObservationRecord.find(verified_query).count()

    pending_query = {**query, "verification_status": {"$nin": ["Verified", "Rejected"]}}
    pending_obs = await ObservationRecord.find(pending_query).count()

    pipeline = [{"$match": query}, {"$group": {"_id": "$species_name"}}, {"$count": "total"}]
    species_res = await ObservationRecord.aggregate(pipeline).to_list()
    total_species = species_res[0]["total"] if species_res else 0

    total_sites = await MonitoringSite.find_all().count()
    total_devices = await SensorDevice.find_all().count()
    total_uploads = await FieldUpload.find_all().count()
    total_users = await User.find_all().count()
    total_predictions = await PredictionRecord.find_all().count()

    return {
        "total_species": total_species,
        "total_observations": total_obs,
        "verified_observations": verified_obs,
        "pending_observations": pending_obs,
        "total_monitoring_sites": total_sites,
        "total_sensor_devices": total_devices,
        "total_uploads": total_uploads,
        "total_users": total_users,
        "total_predictions": total_predictions,
    }


@router.get("/species")
async def get_reports_species(
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
    species: Optional[str] = None,
    monitoring_site_id: Optional[str] = None,
    verification_status: Optional[str] = None,
    observer_id: Optional[str] = None,
    search: Optional[str] = None,
    current_user: User = Depends(get_current_user),
):
    query = build_filter_query(
        start_date, end_date, species, monitoring_site_id, verification_status, observer_id, search
    )

    # Species distribution (top 10)
    dist_pipeline = [
        {"$match": query},
        {"$group": {"_id": "$species_name", "count": {"$sum": 1}}},
        {"$sort": {"count": -1}},
        {"$limit": 10},
        {"$project": {"species": "$_id", "count": 1, "_id": 0}},
    ]
    species_distribution = await ObservationRecord.aggregate(dist_pipeline).to_list()

    # Verification status distribution
    status_pipeline = [
        {"$match": query},
        {"$group": {"_id": "$verification_status", "count": {"$sum": 1}}},
        {"$project": {"status": "$_id", "count": 1, "_id": 0}},
    ]
    verification_distribution = await ObservationRecord.aggregate(status_pipeline).to_list()

    # Monthly observation trend (last 12 months)
    monthly_pipeline = [
        {"$match": query},
        {
            "$group": {
                "_id": {"year": {"$year": "$observed_at"}, "month": {"$month": "$observed_at"}},
                "count": {"$sum": 1},
            }
        },
        {"$sort": {"_id.year": 1, "_id.month": 1}},
        {"$limit": 12},
    ]
    monthly_res = await ObservationRecord.aggregate(monthly_pipeline).to_list()

    months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    monthly_observations = []
    for item in monthly_res:
        m_name = months[item["_id"]["month"] - 1]
        monthly_observations.append({"month": f"{m_name} {item['_id']['year']}", "count": item["count"]})

    return {
        "species_distribution": species_distribution,
        "verification_distribution": verification_distribution,
        "monthly_observations": monthly_observations,
    }


@router.get("/observations")
async def get_reports_observations(
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
    species: Optional[str] = None,
    monitoring_site_id: Optional[str] = None,
    verification_status: Optional[str] = None,
    observer_id: Optional[str] = None,
    search: Optional[str] = None,
    skip: int = Query(0, ge=0),
    limit: int = Query(10, ge=1, le=100),
    sort_by: str = Query("observed_at"),
    sort_order: int = Query(-1),
    current_user: User = Depends(get_current_user),
):
    query = build_filter_query(
        start_date, end_date, species, monitoring_site_id, verification_status, observer_id, search
    )

    total = await ObservationRecord.find(query).count()
    sort_tuple = (sort_by, sort_order)
    observations = await ObservationRecord.find(query).sort(sort_tuple).skip(skip).limit(limit).to_list()

    return {"total": total, "skip": skip, "limit": limit, "observations": observations}


@router.get("/export/excel")
async def export_excel(
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
    species: Optional[str] = None,
    monitoring_site_id: Optional[str] = None,
    verification_status: Optional[str] = None,
    observer_id: Optional[str] = None,
    search: Optional[str] = None,
    current_user: User = Depends(get_current_user),
):
    query = build_filter_query(
        start_date, end_date, species, monitoring_site_id, verification_status, observer_id, search
    )
    observations = await ObservationRecord.find(query).sort("-observed_at").to_list()

    wb = Workbook()

    # --- Styles ---
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="22c55e", end_color="22c55e", fill_type="solid")
    title_font = Font(name="Calibri", bold=True, size=14, color="1a1a1a")
    subtitle_font = Font(name="Calibri", italic=True, size=10, color="6b7280")
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="e5e7eb"),
        right=Side(style="thin", color="e5e7eb"),
        top=Side(style="thin", color="e5e7eb"),
        bottom=Side(style="thin", color="e5e7eb"),
    )

    # ─── Sheet 1: Dashboard Summary ──────────────────────────────────────────
    ws_summary = wb.active
    ws_summary.title = "Dashboard Summary"
    ws_summary.column_dimensions["A"].width = 30
    ws_summary.column_dimensions["B"].width = 20

    ws_summary.merge_cells("A1:B1")
    ws_summary["A1"] = "Wildlife Population Intelligence System"
    ws_summary["A1"].font = title_font
    ws_summary["A1"].alignment = center_align

    ws_summary.merge_cells("A2:B2")
    ws_summary["A2"] = f"Report Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}  |  By: {current_user.full_name}"
    ws_summary["A2"].font = subtitle_font
    ws_summary["A2"].alignment = center_align

    ws_summary.append([])  # blank

    stats = [
        ("Metric", "Value"),
        ("Total Observations", len(observations)),
        ("Verified Observations", sum(1 for o in observations if o.verification_status == "Verified")),
        ("Pending Observations", sum(1 for o in observations if o.verification_status not in ["Verified", "Rejected"])),
        ("Rejected Observations", sum(1 for o in observations if o.verification_status == "Rejected")),
        ("Unique Species", len(set(o.species_name for o in observations))),
    ]
    for i, (label, value) in enumerate(stats):
        row_num = ws_summary.max_row + 1
        ws_summary.append([label, value])
        if i == 0:
            ws_summary[f"A{row_num}"].font = Font(bold=True, color="FFFFFF")
            ws_summary[f"B{row_num}"].font = Font(bold=True, color="FFFFFF")
            ws_summary[f"A{row_num}"].fill = header_fill
            ws_summary[f"B{row_num}"].fill = header_fill
        ws_summary[f"A{row_num}"].border = thin_border
        ws_summary[f"B{row_num}"].border = thin_border

    # ─── Sheet 2: Observation Records ────────────────────────────────────────
    ws_obs = wb.create_sheet(title="Observation Records")
    obs_headers = [
        "Species Name", "Scientific Name", "Observation Type",
        "Confidence (%)", "Monitoring Site", "Observer",
        "Date & Time", "Count", "Verification Status", "Notes",
    ]
    ws_obs.append(obs_headers)
    header_row = ws_obs[1]
    for cell in header_row:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    ws_obs.auto_filter.ref = f"A1:{get_column_letter(len(obs_headers))}1"
    ws_obs.freeze_panes = "A2"

    for obs in observations:
        conf = obs.confidence_score if obs.confidence_score is not None else ""
        date_str = obs.observed_at.strftime("%Y-%m-%d %H:%M")
        row = [
            obs.species_name,
            obs.scientific_name or "",
            obs.observation_type,
            conf,
            obs.monitoring_site_name,
            obs.observer_name,
            date_str,
            obs.count,
            obs.verification_status,
            obs.notes or "",
        ]
        ws_obs.append(row)
        for cell in ws_obs[ws_obs.max_row]:
            cell.border = thin_border

    # Auto-size columns
    obs_col_widths = [22, 22, 18, 14, 22, 20, 18, 8, 20, 30]
    for i, width in enumerate(obs_col_widths, 1):
        ws_obs.column_dimensions[get_column_letter(i)].width = width

    # ─── Sheet 3: Species Statistics ─────────────────────────────────────────
    ws_stats = wb.create_sheet(title="Species Statistics")
    stats_headers = ["Rank", "Species Name", "Observation Count", "% of Total"]
    ws_stats.append(stats_headers)
    for cell in ws_stats[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    ws_stats.auto_filter.ref = f"A1:{get_column_letter(len(stats_headers))}1"
    ws_stats.freeze_panes = "A2"

    species_counts: dict[str, int] = {}
    for obs in observations:
        species_counts[obs.species_name] = species_counts.get(obs.species_name, 0) + 1

    total_count = len(observations) or 1
    for rank, (sp, count) in enumerate(sorted(species_counts.items(), key=lambda x: x[1], reverse=True), 1):
        pct = round((count / total_count) * 100, 1)
        ws_stats.append([rank, sp, count, f"{pct}%"])
        for cell in ws_stats[ws_stats.max_row]:
            cell.border = thin_border

    stats_col_widths = [8, 28, 20, 12]
    for i, width in enumerate(stats_col_widths, 1):
        ws_stats.column_dimensions[get_column_letter(i)].width = width

    # ─── Sheet 4: AI Predictions ─────────────────────────────────────────────
    ws_preds = wb.create_sheet(title="AI Predictions")
    preds_headers = [
        "Prediction ID", "Image Name", "Species Name",
        "Confidence (%)", "Prediction Time (s)", "Model Version",
        "User", "Status", "Date & Time",
    ]
    ws_preds.append(preds_headers)
    for cell in ws_preds[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    ws_preds.auto_filter.ref = f"A1:{get_column_letter(len(preds_headers))}1"
    ws_preds.freeze_panes = "A2"

    predictions = await PredictionRecord.find_all().sort("-created_at").to_list()
    for pred in predictions:
        pred_date = pred.created_at.strftime("%Y-%m-%d %H:%M") if pred.created_at else ""
        row = [
            str(pred.id),
            pred.file_name,
            pred.species_name,
            pred.confidence_score,
            pred.prediction_time,
            pred.model_version,
            pred.user_name,
            pred.status,
            pred_date,
        ]
        ws_preds.append(row)
        for cell in ws_preds[ws_preds.max_row]:
            cell.border = thin_border

    preds_col_widths = [24, 25, 22, 16, 18, 14, 18, 14, 18]
    for i, width in enumerate(preds_col_widths, 1):
        ws_preds.column_dimensions[get_column_letter(i)].width = width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    import asyncio
    asyncio.create_task(
        Notification(
            title="Report Generated",
            message="Excel report has been generated successfully.",
            type="report",
            priority="Low",
            user_id=str(current_user.id)
        ).insert()
    )

    return Response(
        content=output.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="WPIS_Report_{datetime.now().strftime("%Y%m%d")}.xlsx"'},
    )


@router.get("/export/pdf")
async def export_pdf(
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
    species: Optional[str] = None,
    monitoring_site_id: Optional[str] = None,
    verification_status: Optional[str] = None,
    observer_id: Optional[str] = None,
    search: Optional[str] = None,
    current_user: User = Depends(get_current_user),
):
    query = build_filter_query(
        start_date, end_date, species, monitoring_site_id, verification_status, observer_id, search
    )
    observations = await ObservationRecord.find(query).sort("-observed_at").to_list()

    class PDFReport(FPDF):
        def header(self):
            # Green banner
            self.set_fill_color(34, 197, 94)
            self.rect(0, 0, 210, 28, "F")
            self.set_font("helvetica", "B", 16)
            self.set_text_color(255, 255, 255)
            self.set_y(5)
            self.cell(0, 8, "Wildlife Population Intelligence System", align="C", new_x="LMARGIN", new_y="NEXT")
            self.set_font("helvetica", "", 10)
            self.cell(0, 7, "Observation Data Export Report", align="C", new_x="LMARGIN", new_y="NEXT")
            self.set_text_color(30, 30, 30)
            self.ln(5)

        def footer(self):
            self.set_y(-13)
            self.set_draw_color(34, 197, 94)
            self.line(10, self.get_y(), 200, self.get_y())
            self.set_font("helvetica", "I", 8)
            self.set_text_color(120, 120, 120)
            self.cell(0, 8, f"Generated by WPIS  |  Page {self.page_no()}", align="C")

    pdf = PDFReport()
    pdf.set_auto_page_break(auto=True, margin=16)
    pdf.add_page()

    # ── Meta info ──
    pdf.set_font("helvetica", "", 9)
    pdf.set_text_color(80, 80, 80)
    pdf.cell(95, 6, f"Generated At: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", new_x="RIGHT", new_y="TOP")
    pdf.cell(95, 6, f"Generated By: {current_user.full_name}", align="R", new_x="LMARGIN", new_y="NEXT")

    # Active filters
    active_filters = []
    if start_date:
        active_filters.append(f"From: {start_date.strftime('%Y-%m-%d')}")
    if end_date:
        active_filters.append(f"To: {end_date.strftime('%Y-%m-%d')}")
    if species:
        active_filters.append(f"Species: {species}")
    if verification_status:
        active_filters.append(f"Status: {verification_status}")
    if search:
        active_filters.append(f"Search: {search}")

    if active_filters:
        pdf.set_font("helvetica", "I", 9)
        pdf.set_text_color(100, 100, 100)
        pdf.cell(0, 6, "Active Filters: " + "  |  ".join(active_filters), new_x="LMARGIN", new_y="NEXT")

    pdf.ln(4)

    # ── Summary Section ──
    pdf.set_fill_color(240, 253, 244)  # light green bg
    pdf.set_draw_color(187, 247, 208)
    pdf.set_font("helvetica", "B", 11)
    pdf.set_text_color(21, 128, 61)
    pdf.cell(0, 8, "  Summary Statistics", fill=True, border=1, new_x="LMARGIN", new_y="NEXT")

    verified_count = sum(1 for o in observations if o.verification_status == "Verified")
    pending_count = sum(1 for o in observations if o.verification_status not in ["Verified", "Rejected"])
    unique_species = len(set(o.species_name for o in observations))

    stats_data = [
        ("Total Observations", str(len(observations))),
        ("Verified", str(verified_count)),
        ("Pending", str(pending_count)),
        ("Unique Species", str(unique_species)),
    ]
    pdf.set_font("helvetica", "", 10)
    pdf.set_text_color(30, 30, 30)
    col_w = 47
    for i, (label, val) in enumerate(stats_data):
        pdf.set_fill_color(255, 255, 255) if i % 2 == 0 else pdf.set_fill_color(249, 250, 251)
        pdf.cell(col_w, 7, f"  {label}:", fill=True, border=1)
        pdf.cell(col_w, 7, f"  {val}", fill=True, border=1, new_x="RIGHT" if i % 2 == 0 else "LMARGIN", new_y="TOP" if i % 2 == 0 else "NEXT")

    pdf.ln(8)

    # ── Observations Table ──
    pdf.set_font("helvetica", "B", 11)
    pdf.set_fill_color(240, 253, 244)
    pdf.set_draw_color(187, 247, 208)
    pdf.set_text_color(21, 128, 61)
    pdf.cell(0, 8, "  Observation Records", fill=True, border=1, new_x="LMARGIN", new_y="NEXT")
    pdf.ln(2)

    # Table header row
    col_widths = [42, 18, 35, 28, 28, 28]  # Species, Conf, Site, Observer, Date, Status → 179 total
    col_headers = ["Species", "Conf.", "Site", "Observer", "Date", "Status"]
    pdf.set_fill_color(34, 197, 94)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("helvetica", "B", 9)
    for w, h in zip(col_widths, col_headers):
        pdf.cell(w, 9, f"  {h}", border=1, fill=True)
    pdf.ln()

    pdf.set_font("helvetica", "", 8)
    for row_idx, obs in enumerate(observations):
        # Alternate row shading
        if row_idx % 2 == 0:
            pdf.set_fill_color(255, 255, 255)
        else:
            pdf.set_fill_color(249, 250, 251)
        pdf.set_text_color(30, 30, 30)

        conf = f"{obs.confidence_score}%" if obs.confidence_score is not None else "N/A"
        date_str = obs.observed_at.strftime("%Y-%m-%d")

        def trunc(s, n):
            return (s[: n - 2] + "..") if len(s) > n else s

        values = [
            trunc(obs.species_name, 22),
            conf,
            trunc(obs.monitoring_site_name or "", 18),
            trunc(obs.observer_name or "", 15),
            date_str,
            trunc(obs.verification_status, 15),
        ]
        for w, v in zip(col_widths, values):
            pdf.cell(w, 7, f"  {v}", border=1, fill=True)
        pdf.ln()

    pdf.ln(8)
    
    # ── AI Predictions Table ──
    pdf.set_font("helvetica", "B", 11)
    pdf.set_fill_color(240, 253, 244)
    pdf.set_draw_color(187, 247, 208)
    pdf.set_text_color(21, 128, 61)
    pdf.cell(0, 8, "  AI Prediction Records (Recent)", fill=True, border=1, new_x="LMARGIN", new_y="NEXT")
    pdf.ln(2)

    col_widths_pred = [45, 20, 25, 30, 28, 31]  # Species, Conf, Time, User, Date, Status
    col_headers_pred = ["Species", "Conf.", "Time (s)", "User", "Date", "Status"]
    pdf.set_fill_color(34, 197, 94)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("helvetica", "B", 9)
    for w, h in zip(col_widths_pred, col_headers_pred):
        pdf.cell(w, 9, f"  {h}", border=1, fill=True)
    pdf.ln()

    pdf.set_font("helvetica", "", 8)
    pdf_predictions = await PredictionRecord.find_all().sort("-created_at").limit(50).to_list()
    for row_idx, pred in enumerate(pdf_predictions):
        if row_idx % 2 == 0:
            pdf.set_fill_color(255, 255, 255)
        else:
            pdf.set_fill_color(249, 250, 251)
        pdf.set_text_color(30, 30, 30)

        conf = f"{pred.confidence_score}%"
        date_str = pred.created_at.strftime("%Y-%m-%d") if pred.created_at else ""

        def trunc_s(s, n):
            return (s[: n - 2] + "..") if len(s) > n else s

        values = [
            trunc_s(pred.species_name, 22),
            conf,
            f"{pred.prediction_time}s",
            trunc_s(pred.user_name or "", 15),
            date_str,
            trunc_s(pred.status, 15),
        ]
        for w, v in zip(col_widths_pred, values):
            pdf.cell(w, 7, f"  {v}", border=1, fill=True)
        pdf.ln()

    pdf_bytes = pdf.output(dest="S")
    
    import asyncio
    asyncio.create_task(
        Notification(
            title="Report Generated",
            message="PDF report has been generated successfully.",
            type="report",
            priority="Low",
            user_id=str(current_user.id)
        ).insert()
    )

    return Response(
        content=bytes(pdf_bytes),
        media_type="application/pdf",
        headers={
            "Content-Disposition": f'attachment; filename="WPIS_Report_{datetime.now().strftime("%Y%m%d")}.pdf"'
        },
    )
