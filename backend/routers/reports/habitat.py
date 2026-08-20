from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from io import BytesIO, StringIO
import csv
import re
import html
from urllib.parse import quote

from openpyxl import Workbook

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
    PageBreak,
)

import models
from database import get_db
from auth import get_current_user


router = APIRouter(
    prefix="/reports",
    tags=["Habitat Assessment Reports"]
)


# =========================================================
# HELPERS
# =========================================================

def safe_value(value, default="-"):
    if value is None:
        return default

    value = str(value).strip()

    if value == "":
        return default

    return value


def safe_pdf_text(value, default="-"):
    return html.escape(
        safe_value(value, default),
        quote=True
    )


def safe_excel_sheet_name(name, fallback="Habitat"):
    if not name:
        name = fallback

    name = str(name)

    name = re.sub(
        r'[:\\/?*\[\]]',
        "_",
        name
    )

    name = name.strip()

    if not name:
        name = fallback

    return name[:31]


# =========================================================
# BUILD SINGLE HABITAT REPORT
# =========================================================

def build_habitat_report(
    db: Session,
    habitat_name: str
):

    species_list = (
        db.query(models.Species)
        .filter(
            models.Species.habitat == habitat_name
        )
        .order_by(
            models.Species.species_name.asc()
        )
        .all()
    )

    species_data = []

    total_population = 0
    total_observations = 0
    total_observed_count = 0

    locations = set()
    conservation_statuses = set()
    categories = set()

    for species in species_list:

        if species.category:
            categories.add(
                species.category
            )

        if species.conservation_status:
            conservation_statuses.add(
                species.conservation_status
            )

        population = (
            species.population or 0
        )

        total_population += population

        observations = (
            db.query(models.Observation)
            .filter(
                models.Observation.species_id
                == species.id
            )
            .order_by(
                models.Observation.observation_date.desc()
            )
            .all()
        )

        observation_data = []

        species_observed_count = 0
        species_locations = set()

        for observation in observations:

            if observation.location:
                locations.add(
                    observation.location
                )

                species_locations.add(
                    observation.location
                )

            population_count = (
                observation.population_count
                if observation.population_count is not None
                else 0
            )

            species_observed_count += (
                population_count
            )

            total_observed_count += (
                population_count
            )

            observation_data.append({
                "id": observation.id,
                "survey_id": observation.survey_id,
                "location": observation.location,
                "latitude": observation.latitude,
                "longitude": observation.longitude,
                "observation_date":
                    observation.observation_date,
                "observer_name":
                    observation.observer_name,
                "population_count":
                    observation.population_count,
                "image_path":
                    observation.image_path,
                "audio_path":
                    observation.audio_path,
                "notes":
                    observation.notes,
            })

        total_observations += len(
            observations
        )

        species_data.append({
            "species": {
                "id": species.id,
                "species_name":
                    species.species_name,
                "scientific_name":
                    species.scientific_name,
                "category":
                    species.category,
                "population":
                    species.population,
                "conservation_status":
                    species.conservation_status,
                "habitat":
                    species.habitat,
                "image":
                    species.image,
            },

            "summary": {
                "population":
                    population,

                "total_observations":
                    len(observations),

                "total_observed_count":
                    species_observed_count,

                "locations":
                    len(species_locations),
            },

            "observations":
                observation_data,
        })

    return {
        "report_type":
            "Habitat Assessment Report",

        "habitat": {
            "name":
                habitat_name,
        },

        "summary": {
            "habitat":
                habitat_name,

            "total_species":
                len(species_list),

            "total_population":
                total_population,

            "total_observations":
                total_observations,

            "total_observed_count":
                total_observed_count,

            "locations":
                len(locations),

            "categories":
                len(categories),

            "conservation_statuses":
                len(conservation_statuses),
        },

        "analysis": {
            "categories":
                sorted(categories),

            "conservation_statuses":
                sorted(
                    conservation_statuses
                ),

            "locations":
                sorted(locations),
        },

        "species":
            species_data,
    }


# =========================================================
# GET ALL HABITAT ASSESSMENT REPORT
# =========================================================

@router.get("/habitat")
def get_habitat_report(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):

    species_list = (
        db.query(models.Species)
        .order_by(
            models.Species.species_name.asc()
        )
        .all()
    )

    habitats = set()

    for species in species_list:

        if species.habitat:
            habitat_name = str(
                species.habitat
            ).strip()

            if habitat_name:
                habitats.add(
                    habitat_name
                )

    habitat_reports = []

    total_population = 0
    total_observations = 0
    total_observed_count = 0
    locations = set()

    for habitat_name in sorted(habitats):

        report = build_habitat_report(
            db,
            habitat_name
        )

        habitat_reports.append(
            report
        )

        summary = report["summary"]

        total_population += (
            summary["total_population"]
        )

        total_observations += (
            summary["total_observations"]
        )

        total_observed_count += (
            summary["total_observed_count"]
        )

        for location in report[
            "analysis"
        ]["locations"]:

            locations.add(location)

    return {
        "report_type":
            "Habitat Assessment Report",

        "summary": {
            "total_habitats":
                len(habitat_reports),

            "total_species":
                len(species_list),

            "total_population":
                total_population,

            "total_observations":
                total_observations,

            "total_observed_count":
                total_observed_count,

            "locations":
                len(locations),
        },

        "habitats":
            habitat_reports,
    }


# =========================================================
# GET SINGLE HABITAT REPORT
# =========================================================

@router.get("/habitat/{habitat_name}")
def get_single_habitat_report(
    habitat_name: str,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):

    decoded_habitat = habitat_name.strip()

    species_exists = (
        db.query(models.Species)
        .filter(
            models.Species.habitat
            == decoded_habitat
        )
        .first()
    )

    if not species_exists:

        raise HTTPException(
            status_code=404,
            detail="Habitat not found"
        )

    return build_habitat_report(
        db,
        decoded_habitat
    )


# =========================================================
# EXPORT ALL HABITATS - PDF
# =========================================================

@router.get("/habitat/all/pdf")
def export_all_habitat_pdf(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):

    report = get_habitat_report(
        db,
        current_user
    )

    buffer = BytesIO()

    document = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=30,
        leftMargin=30,
        topMargin=30,
        bottomMargin=30,
    )

    styles = getSampleStyleSheet()

    story = []

    story.append(
        Paragraph(
            "Habitat Assessment Report",
            styles["Title"]
        )
    )

    story.append(
        Spacer(1, 15)
    )

    summary = report["summary"]

    summary_data = [
        ["Metric", "Value"],

        [
            "Total Habitats",
            safe_value(
                summary["total_habitats"]
            )
        ],

        [
            "Total Species",
            safe_value(
                summary["total_species"]
            )
        ],

        [
            "Total Population",
            safe_value(
                summary["total_population"]
            )
        ],

        [
            "Total Observations",
            safe_value(
                summary["total_observations"]
            )
        ],

        [
            "Total Observed Count",
            safe_value(
                summary["total_observed_count"]
            )
        ],

        [
            "Locations",
            safe_value(
                summary["locations"]
            )
        ],
    ]

    summary_table = Table(
        summary_data,
        colWidths=[220, 120]
    )

    summary_table.setStyle(
        TableStyle([
            (
                "BACKGROUND",
                (0, 0),
                (-1, 0),
                colors.HexColor("#6f42c1")
            ),

            (
                "TEXTCOLOR",
                (0, 0),
                (-1, 0),
                colors.white
            ),

            (
                "FONTNAME",
                (0, 0),
                (-1, 0),
                "Helvetica-Bold"
            ),

            (
                "GRID",
                (0, 0),
                (-1, -1),
                0.5,
                colors.grey
            ),

            (
                "PADDING",
                (0, 0),
                (-1, -1),
                7
            ),
        ])
    )

    story.append(
        summary_table
    )

    story.append(
        Spacer(1, 20)
    )

    for habitat_report in report["habitats"]:

        habitat = habitat_report[
            "habitat"
        ]["name"]

        summary = habitat_report[
            "summary"
        ]

        story.append(
            Paragraph(
                f"Habitat: {safe_pdf_text(habitat)}",
                styles["Heading2"]
            )
        )

        habitat_summary = [
            ["Metric", "Value"],

            [
                "Species",
                safe_value(
                    summary["total_species"]
                )
            ],

            [
                "Population",
                safe_value(
                    summary["total_population"]
                )
            ],

            [
                "Observations",
                safe_value(
                    summary["total_observations"]
                )
            ],

            [
                "Observed Count",
                safe_value(
                    summary["total_observed_count"]
                )
            ],

            [
                "Locations",
                safe_value(
                    summary["locations"]
                )
            ],
        ]

        table = Table(
            habitat_summary,
            colWidths=[220, 120]
        )

        table.setStyle(
            TableStyle([
                (
                    "BACKGROUND",
                    (0, 0),
                    (-1, 0),
                    colors.HexColor("#6f42c1")
                ),

                (
                    "TEXTCOLOR",
                    (0, 0),
                    (-1, 0),
                    colors.white
                ),

                (
                    "GRID",
                    (0, 0),
                    (-1, -1),
                    0.5,
                    colors.grey
                ),

                (
                    "PADDING",
                    (0, 0),
                    (-1, -1),
                    6
                ),
            ])
        )

        story.append(table)

        story.append(
            Spacer(1, 10)
        )

        species_table_data = [
            [
                "Species",
                "Scientific Name",
                "Category",
                "Population",
                "Conservation Status",
                "Observations",
            ]
        ]

        for species_data in habitat_report[
            "species"
        ]:

            species = species_data[
                "species"
            ]

            species_summary = (
                species_data["summary"]
            )

            species_table_data.append([
                safe_pdf_text(
                    species["species_name"]
                ),

                safe_pdf_text(
                    species["scientific_name"]
                ),

                safe_pdf_text(
                    species["category"]
                ),

                safe_value(
                    species["population"]
                ),

                safe_pdf_text(
                    species[
                        "conservation_status"
                    ]
                ),

                safe_value(
                    species_summary[
                        "total_observations"
                    ]
                ),
            ])

        species_table = Table(
            species_table_data,
            repeatRows=1,
        )

        species_table.setStyle(
            TableStyle([
                (
                    "BACKGROUND",
                    (0, 0),
                    (-1, 0),
                    colors.HexColor("#432874")
                ),

                (
                    "TEXTCOLOR",
                    (0, 0),
                    (-1, 0),
                    colors.white
                ),

                (
                    "FONTNAME",
                    (0, 0),
                    (-1, 0),
                    "Helvetica-Bold"
                ),

                (
                    "GRID",
                    (0, 0),
                    (-1, -1),
                    0.5,
                    colors.grey
                ),

                (
                    "VALIGN",
                    (0, 0),
                    (-1, -1),
                    "TOP"
                ),

                (
                    "PADDING",
                    (0, 0),
                    (-1, -1),
                    5
                ),
            ])
        )

        story.append(
            species_table
        )

        story.append(
            Spacer(1, 20)
        )

    document.build(story)

    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={
            "Content-Disposition":
                "attachment; filename=all-habitat-assessment-reports.pdf"
        },
    )


# =========================================================
# EXPORT ALL HABITATS - EXCEL
# =========================================================

@router.get("/habitat/all/excel")
def export_all_habitat_excel(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):

    report = get_habitat_report(
        db,
        current_user
    )

    workbook = Workbook()

    default_sheet = workbook.active
    workbook.remove(default_sheet)

    summary_sheet = workbook.create_sheet(
        "Summary"
    )

    summary_sheet.append([
        "Habitat Assessment Report"
    ])

    summary_sheet.append([])

    summary_sheet.append([
        "Metric",
        "Value"
    ])

    summary = report["summary"]

    summary_sheet.append([
        "Total Habitats",
        summary["total_habitats"]
    ])

    summary_sheet.append([
        "Total Species",
        summary["total_species"]
    ])

    summary_sheet.append([
        "Total Population",
        summary["total_population"]
    ])

    summary_sheet.append([
        "Total Observations",
        summary["total_observations"]
    ])

    summary_sheet.append([
        "Total Observed Count",
        summary["total_observed_count"]
    ])

    summary_sheet.append([
        "Locations",
        summary["locations"]
    ])

    for cell in summary_sheet[1]:
        cell.font = cell.font.copy(
            bold=True
        )

    for habitat_report in report["habitats"]:

        habitat_name = habitat_report[
            "habitat"
        ]["name"]

        sheet_name = safe_excel_sheet_name(
            habitat_name,
            "Habitat"
        )

        sheet = workbook.create_sheet(
            sheet_name
        )

        habitat_summary = (
            habitat_report["summary"]
        )

        sheet.append([
            "Habitat",
            habitat_name
        ])

        sheet.append([
            "Total Species",
            habitat_summary[
                "total_species"
            ]
        ])

        sheet.append([
            "Total Population",
            habitat_summary[
                "total_population"
            ]
        ])

        sheet.append([
            "Total Observations",
            habitat_summary[
                "total_observations"
            ]
        ])

        sheet.append([
            "Total Observed Count",
            habitat_summary[
                "total_observed_count"
            ]
        ])

        sheet.append([
            "Locations",
            habitat_summary[
                "locations"
            ]
        ])

        sheet.append([])

        sheet.append([
            "Species",
            "Scientific Name",
            "Category",
            "Population",
            "Conservation Status",
            "Observations",
            "Observed Count",
            "Locations",
        ])

        for species_data in habitat_report[
            "species"
        ]:

            species = species_data[
                "species"
            ]

            species_summary = (
                species_data["summary"]
            )

            sheet.append([
                species["species_name"],
                species["scientific_name"],
                species["category"],
                species["population"],
                species[
                    "conservation_status"
                ],
                species_summary[
                    "total_observations"
                ],
                species_summary[
                    "total_observed_count"
                ],
                species_summary[
                    "locations"
                ],
            ])

        for cell in sheet[1]:
            cell.font = cell.font.copy(
                bold=True
            )

        for cell in sheet[8]:
            cell.font = cell.font.copy(
                bold=True
            )

        sheet.freeze_panes = "A9"

    output = BytesIO()

    workbook.save(output)

    output.seek(0)

    return StreamingResponse(
        output,
        media_type=(
            "application/vnd.openxmlformats-"
            "officedocument.spreadsheetml.sheet"
        ),
        headers={
            "Content-Disposition":
                "attachment; filename=all-habitat-assessment-reports.xlsx"
        },
    )


# =========================================================
# EXPORT ALL HABITATS - CSV
# =========================================================

@router.get("/habitat/all/csv")
def export_all_habitat_csv(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):

    report = get_habitat_report(
        db,
        current_user
    )

    output = StringIO()

    writer = csv.writer(output)

    writer.writerow([
        "Habitat",
        "Species",
        "Scientific Name",
        "Category",
        "Population",
        "Conservation Status",
        "Observations",
        "Observed Count",
        "Locations",
    ])

    for habitat_report in report["habitats"]:

        habitat_name = habitat_report[
            "habitat"
        ]["name"]

        for species_data in habitat_report[
            "species"
        ]:

            species = species_data[
                "species"
            ]

            summary = species_data[
                "summary"
            ]

            writer.writerow([
                habitat_name,
                safe_value(
                    species["species_name"]
                ),
                safe_value(
                    species["scientific_name"]
                ),
                safe_value(
                    species["category"]
                ),
                safe_value(
                    species["population"]
                ),
                safe_value(
                    species[
                        "conservation_status"
                    ]
                ),
                safe_value(
                    summary[
                        "total_observations"
                    ]
                ),
                safe_value(
                    summary[
                        "total_observed_count"
                    ]
                ),
                safe_value(
                    summary[
                        "locations"
                    ]
                ),
            ])

    output.seek(0)

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={
            "Content-Disposition":
                "attachment; filename=all-habitat-assessment-reports.csv"
        },
    )


# =========================================================
# SINGLE HABITAT - PDF
# =========================================================

@router.get("/habitat/{habitat_name}/pdf")
def export_single_habitat_pdf(
    habitat_name: str,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):

    report = get_single_habitat_report(
        habitat_name,
        db,
        current_user
    )

    buffer = BytesIO()

    document = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=30,
        leftMargin=30,
        topMargin=30,
        bottomMargin=30,
    )

    styles = getSampleStyleSheet()

    story = []

    habitat = report["habitat"]["name"]
    summary = report["summary"]

    story.append(
        Paragraph(
            f"Habitat Assessment Report - "
            f"{safe_pdf_text(habitat)}",
            styles["Title"]
        )
    )

    story.append(
        Spacer(1, 15)
    )

    summary_data = [
        ["Metric", "Value"],

        [
            "Habitat",
            safe_pdf_text(habitat)
        ],

        [
            "Total Species",
            safe_value(
                summary["total_species"]
            )
        ],

        [
            "Total Population",
            safe_value(
                summary["total_population"]
            )
        ],

        [
            "Total Observations",
            safe_value(
                summary["total_observations"]
            )
        ],

        [
            "Total Observed Count",
            safe_value(
                summary["total_observed_count"]
            )
        ],

        [
            "Locations",
            safe_value(
                summary["locations"]
            )
        ],
    ]

    table = Table(
        summary_data,
        colWidths=[220, 200]
    )

    table.setStyle(
        TableStyle([
            (
                "BACKGROUND",
                (0, 0),
                (-1, 0),
                colors.HexColor("#6f42c1")
            ),

            (
                "TEXTCOLOR",
                (0, 0),
                (-1, 0),
                colors.white
            ),

            (
                "GRID",
                (0, 0),
                (-1, -1),
                0.5,
                colors.grey
            ),

            (
                "PADDING",
                (0, 0),
                (-1, -1),
                7
            ),
        ])
    )

    story.append(table)

    story.append(
        Spacer(1, 20)
    )

    story.append(
        Paragraph(
            "Species in this Habitat",
            styles["Heading2"]
        )
    )

    species_table_data = [
        [
            "Species",
            "Scientific Name",
            "Category",
            "Population",
            "Conservation Status",
            "Observations",
        ]
    ]

    for species_data in report["species"]:

        species = species_data[
            "species"
        ]

        species_summary = species_data[
            "summary"
        ]

        species_table_data.append([
            safe_pdf_text(
                species["species_name"]
            ),

            safe_pdf_text(
                species["scientific_name"]
            ),

            safe_pdf_text(
                species["category"]
            ),

            safe_value(
                species["population"]
            ),

            safe_pdf_text(
                species[
                    "conservation_status"
                ]
            ),

            safe_value(
                species_summary[
                    "total_observations"
                ]
            ),
        ])

    species_table = Table(
        species_table_data,
        repeatRows=1,
    )

    species_table.setStyle(
        TableStyle([
            (
                "BACKGROUND",
                (0, 0),
                (-1, 0),
                colors.HexColor("#432874")
            ),

            (
                "TEXTCOLOR",
                (0, 0),
                (-1, 0),
                colors.white
            ),

            (
                "FONTNAME",
                (0, 0),
                (-1, 0),
                "Helvetica-Bold"
            ),

            (
                "GRID",
                (0, 0),
                (-1, -1),
                0.5,
                colors.grey
            ),

            (
                "PADDING",
                (0, 0),
                (-1, -1),
                5
            ),
        ])
    )

    story.append(
        species_table
    )

    document.build(story)

    buffer.seek(0)

    filename = (
        f"{habitat_name}"
        "-habitat-assessment-report.pdf"
    )

    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={
            "Content-Disposition":
                f'attachment; filename="{filename}"'
        },
    )


# =========================================================
# SINGLE HABITAT - EXCEL
# =========================================================

@router.get("/habitat/{habitat_name}/excel")
def export_single_habitat_excel(
    habitat_name: str,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):

    report = get_single_habitat_report(
        habitat_name,
        db,
        current_user
    )

    workbook = Workbook()

    sheet = workbook.active

    sheet.title = safe_excel_sheet_name(
        habitat_name
    )

    summary = report["summary"]

    sheet.append([
        "Habitat Assessment Report"
    ])

    sheet.append([])

    sheet.append([
        "Habitat",
        habitat_name
    ])

    sheet.append([
        "Total Species",
        summary["total_species"]
    ])

    sheet.append([
        "Total Population",
        summary["total_population"]
    ])

    sheet.append([
        "Total Observations",
        summary["total_observations"]
    ])

    sheet.append([
        "Total Observed Count",
        summary["total_observed_count"]
    ])

    sheet.append([
        "Locations",
        summary["locations"]
    ])

    sheet.append([])

    sheet.append([
        "Species",
        "Scientific Name",
        "Category",
        "Population",
        "Conservation Status",
        "Observations",
        "Observed Count",
        "Locations",
    ])

    for species_data in report["species"]:

        species = species_data[
            "species"
        ]

        species_summary = species_data[
            "summary"
        ]

        sheet.append([
            species["species_name"],
            species["scientific_name"],
            species["category"],
            species["population"],
            species[
                "conservation_status"
            ],
            species_summary[
                "total_observations"
            ],
            species_summary[
                "total_observed_count"
            ],
            species_summary[
                "locations"
            ],
        ])

    for cell in sheet[1]:
        cell.font = cell.font.copy(
            bold=True
        )

    for cell in sheet[10]:
        cell.font = cell.font.copy(
            bold=True
        )

    sheet.freeze_panes = "A11"

    output = BytesIO()

    workbook.save(output)

    output.seek(0)

    filename = (
        f"{habitat_name}"
        "-habitat-assessment-report.xlsx"
    )

    return StreamingResponse(
        output,
        media_type=(
            "application/vnd.openxmlformats-"
            "officedocument.spreadsheetml.sheet"
        ),
        headers={
            "Content-Disposition":
                f'attachment; filename="{filename}"'
        },
    )


# =========================================================
# SINGLE HABITAT - CSV
# =========================================================

@router.get("/habitat/{habitat_name}/csv")
def export_single_habitat_csv(
    habitat_name: str,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):

    report = get_single_habitat_report(
        habitat_name,
        db,
        current_user
    )

    output = StringIO()

    writer = csv.writer(output)

    writer.writerow([
        "Habitat",
        "Species",
        "Scientific Name",
        "Category",
        "Population",
        "Conservation Status",
        "Observations",
        "Observed Count",
        "Locations",
    ])

    for species_data in report["species"]:

        species = species_data[
            "species"
        ]

        summary = species_data[
            "summary"
        ]

        writer.writerow([
            habitat_name,
            safe_value(
                species["species_name"]
            ),
            safe_value(
                species["scientific_name"]
            ),
            safe_value(
                species["category"]
            ),
            safe_value(
                species["population"]
            ),
            safe_value(
                species[
                    "conservation_status"
                ]
            ),
            safe_value(
                summary[
                    "total_observations"
                ]
            ),
            safe_value(
                summary[
                    "total_observed_count"
                ]
            ),
            safe_value(
                summary[
                    "locations"
                ]
            ),
        ])

    output.seek(0)

    filename = (
        f"{habitat_name}"
        "-habitat-assessment-report.csv"
    )

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={
            "Content-Disposition":
                f'attachment; filename="{filename}"'
        },
    )