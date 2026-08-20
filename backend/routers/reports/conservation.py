from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from io import BytesIO, StringIO
import csv
import re
import html

from openpyxl import Workbook

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
)

import models
from database import get_db
from auth import get_current_user


router = APIRouter(
    prefix="/reports",
    tags=["Conservation Reports"]
)


# =========================================================
# HELPERS
# =========================================================

def safe_value(value, default="-"):
    if value is None:
        return default

    value = str(value).strip()

    if value == "":
        return default

    return value


def safe_pdf_text(value, default="-"):
    return html.escape(
        safe_value(value, default),
        quote=True
    )


def safe_excel_sheet_name(
    name,
    fallback="Conservation"
):
    if not name:
        name = fallback

    name = str(name)

    name = re.sub(
        r'[:\\/?*\[\]]',
        "_",
        name
    )

    name = name.strip()

    if not name:
        name = fallback

    return name[:31]


# =========================================================
# BUILD SINGLE CONSERVATION STATUS REPORT
# =========================================================

def build_conservation_report(
    db: Session,
    conservation_status: str
):

    species_list = (
        db.query(models.Species)
        .filter(
            models.Species.conservation_status
            == conservation_status
        )
        .order_by(
            models.Species.species_name.asc()
        )
        .all()
    )

    species_data = []

    total_population = 0
    total_observations = 0
    total_observed_count = 0

    locations = set()
    categories = set()
    habitats = set()

    for species in species_list:

        if species.category:
            categories.add(
                species.category
            )

        if species.habitat:
            habitats.add(
                species.habitat
            )

        population = (
            species.population or 0
        )

        total_population += population

        observations = (
            db.query(models.Observation)
            .filter(
                models.Observation.species_id
                == species.id
            )
            .order_by(
                models.Observation.observation_date.desc()
            )
            .all()
        )

        observation_data = []

        species_observed_count = 0
        species_locations = set()

        for observation in observations:

            if observation.location:

                locations.add(
                    observation.location
                )

                species_locations.add(
                    observation.location
                )

            population_count = (
                observation.population_count
                if observation.population_count is not None
                else 0
            )

            species_observed_count += (
                population_count
            )

            total_observed_count += (
                population_count
            )

            observation_data.append({
                "id": observation.id,
                "survey_id": observation.survey_id,
                "location": observation.location,
                "latitude": observation.latitude,
                "longitude": observation.longitude,
                "observation_date":
                    observation.observation_date,
                "observer_name":
                    observation.observer_name,
                "population_count":
                    observation.population_count,
                "image_path":
                    observation.image_path,
                "audio_path":
                    observation.audio_path,
                "notes":
                    observation.notes,
            })

        total_observations += len(
            observations
        )

        species_data.append({
            "species": {
                "id": species.id,
                "species_name":
                    species.species_name,
                "scientific_name":
                    species.scientific_name,
                "category":
                    species.category,
                "population":
                    species.population,
                "conservation_status":
                    species.conservation_status,
                "habitat":
                    species.habitat,
                "image":
                    species.image,
            },

            "summary": {
                "population":
                    population,

                "total_observations":
                    len(observations),

                "total_observed_count":
                    species_observed_count,

                "locations":
                    len(species_locations),
            },

            "observations":
                observation_data,
        })

    return {
        "report_type":
            "Conservation Report",

        "conservation": {
            "status":
                conservation_status,
        },

        "summary": {
            "conservation_status":
                conservation_status,

            "total_species":
                len(species_list),

            "total_population":
                total_population,

            "total_observations":
                total_observations,

            "total_observed_count":
                total_observed_count,

            "locations":
                len(locations),

            "categories":
                len(categories),

            "habitats":
                len(habitats),
        },

        "analysis": {
            "categories":
                sorted(categories),

            "habitats":
                sorted(habitats),

            "locations":
                sorted(locations),
        },

        "species":
            species_data,
    }


# =========================================================
# GET ALL CONSERVATION REPORTS
# =========================================================

@router.get("/conservation")
def get_conservation_report(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):

    species_list = (
        db.query(models.Species)
        .order_by(
            models.Species.species_name.asc()
        )
        .all()
    )

    statuses = set()

    for species in species_list:

        if species.conservation_status:

            status = str(
                species.conservation_status
            ).strip()

            if status:
                statuses.add(status)

    conservation_reports = []

    total_population = 0
    total_observations = 0
    total_observed_count = 0

    locations = set()

    for status in sorted(statuses):

        report = build_conservation_report(
            db,
            status
        )

        conservation_reports.append(
            report
        )

        summary = report["summary"]

        total_population += (
            summary["total_population"]
        )

        total_observations += (
            summary["total_observations"]
        )

        total_observed_count += (
            summary["total_observed_count"]
        )

        for location in report[
            "analysis"
        ]["locations"]:

            locations.add(location)

    return {
        "report_type":
            "Conservation Report",

        "summary": {
            "total_statuses":
                len(conservation_reports),

            "total_species":
                len(species_list),

            "total_population":
                total_population,

            "total_observations":
                total_observations,

            "total_observed_count":
                total_observed_count,

            "locations":
                len(locations),
        },

        "conservation_statuses":
            conservation_reports,
    }


# =========================================================
# GET SINGLE CONSERVATION STATUS REPORT
# =========================================================

@router.get(
    "/conservation/{conservation_status}"
)
def get_single_conservation_report(
    conservation_status: str,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):

    decoded_status = (
        conservation_status.strip()
    )

    species_exists = (
        db.query(models.Species)
        .filter(
            models.Species.conservation_status
            == decoded_status
        )
        .first()
    )

    if not species_exists:

        raise HTTPException(
            status_code=404,
            detail="Conservation status not found"
        )

    return build_conservation_report(
        db,
        decoded_status
    )


# =========================================================
# EXPORT ALL CONSERVATION REPORTS - PDF
# =========================================================

@router.get("/conservation/all/pdf")
def export_all_conservation_pdf(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):

    report = get_conservation_report(
        db,
        current_user
    )

    buffer = BytesIO()

    document = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=30,
        leftMargin=30,
        topMargin=30,
        bottomMargin=30,
    )

    styles = getSampleStyleSheet()

    story = []

    story.append(
        Paragraph(
            "Conservation Report",
            styles["Title"]
        )
    )

    story.append(
        Spacer(1, 15)
    )

    summary = report["summary"]

    summary_data = [
        ["Metric", "Value"],

        [
            "Total Conservation Statuses",
            safe_value(
                summary["total_statuses"]
            )
        ],

        [
            "Total Species",
            safe_value(
                summary["total_species"]
            )
        ],

        [
            "Total Population",
            safe_value(
                summary["total_population"]
            )
        ],

        [
            "Total Observations",
            safe_value(
                summary["total_observations"]
            )
        ],

        [
            "Total Observed Count",
            safe_value(
                summary["total_observed_count"]
            )
        ],

        [
            "Locations",
            safe_value(
                summary["locations"]
            )
        ],
    ]

    summary_table = Table(
        summary_data,
        colWidths=[240, 120]
    )

    summary_table.setStyle(
        TableStyle([
            (
                "BACKGROUND",
                (0, 0),
                (-1, 0),
                colors.HexColor("#dc3545")
            ),

            (
                "TEXTCOLOR",
                (0, 0),
                (-1, 0),
                colors.white
            ),

            (
                "FONTNAME",
                (0, 0),
                (-1, 0),
                "Helvetica-Bold"
            ),

            (
                "GRID",
                (0, 0),
                (-1, -1),
                0.5,
                colors.grey
            ),

            (
                "PADDING",
                (0, 0),
                (-1, -1),
                7
            ),
        ])
    )

    story.append(
        summary_table
    )

    story.append(
        Spacer(1, 20)
    )

    for status_report in report[
        "conservation_statuses"
    ]:

        status = status_report[
            "conservation"
        ]["status"]

        status_summary = status_report[
            "summary"
        ]

        story.append(
            Paragraph(
                f"Conservation Status: "
                f"{safe_pdf_text(status)}",
                styles["Heading2"]
            )
        )

        status_summary_data = [
            ["Metric", "Value"],

            [
                "Species",
                safe_value(
                    status_summary[
                        "total_species"
                    ]
                )
            ],

            [
                "Population",
                safe_value(
                    status_summary[
                        "total_population"
                    ]
                )
            ],

            [
                "Observations",
                safe_value(
                    status_summary[
                        "total_observations"
                    ]
                )
            ],

            [
                "Observed Count",
                safe_value(
                    status_summary[
                        "total_observed_count"
                    ]
                )
            ],

            [
                "Locations",
                safe_value(
                    status_summary[
                        "locations"
                    ]
                )
            ],
        ]

        status_table = Table(
            status_summary_data,
            colWidths=[220, 120]
        )

        status_table.setStyle(
            TableStyle([
                (
                    "BACKGROUND",
                    (0, 0),
                    (-1, 0),
                    colors.HexColor("#dc3545")
                ),

                (
                    "TEXTCOLOR",
                    (0, 0),
                    (-1, 0),
                    colors.white
                ),

                (
                    "GRID",
                    (0, 0),
                    (-1, -1),
                    0.5,
                    colors.grey
                ),

                (
                    "PADDING",
                    (0, 0),
                    (-1, -1),
                    6
                ),
            ])
        )

        story.append(
            status_table
        )

        story.append(
            Spacer(1, 10)
        )

        species_table_data = [
            [
                "Species",
                "Scientific Name",
                "Category",
                "Habitat",
                "Population",
                "Observations",
            ]
        ]

        for species_data in status_report[
            "species"
        ]:

            species = species_data[
                "species"
            ]

            species_summary = (
                species_data["summary"]
            )

            species_table_data.append([
                safe_pdf_text(
                    species["species_name"]
                ),

                safe_pdf_text(
                    species["scientific_name"]
                ),

                safe_pdf_text(
                    species["category"]
                ),

                safe_pdf_text(
                    species["habitat"]
                ),

                safe_value(
                    species["population"]
                ),

                safe_value(
                    species_summary[
                        "total_observations"
                    ]
                ),
            ])

        species_table = Table(
            species_table_data,
            repeatRows=1,
        )

        species_table.setStyle(
            TableStyle([
                (
                    "BACKGROUND",
                    (0, 0),
                    (-1, 0),
                    colors.HexColor("#a71d2a")
                ),

                (
                    "TEXTCOLOR",
                    (0, 0),
                    (-1, 0),
                    colors.white
                ),

                (
                    "FONTNAME",
                    (0, 0),
                    (-1, 0),
                    "Helvetica-Bold"
                ),

                (
                    "GRID",
                    (0, 0),
                    (-1, -1),
                    0.5,
                    colors.grey
                ),

                (
                    "VALIGN",
                    (0, 0),
                    (-1, -1),
                    "TOP"
                ),

                (
                    "PADDING",
                    (0, 0),
                    (-1, -1),
                    5
                ),
            ])
        )

        story.append(
            species_table
        )

        story.append(
            Spacer(1, 20)
        )

    document.build(story)

    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={
            "Content-Disposition":
                "attachment; filename=all-conservation-reports.pdf"
        },
    )


# =========================================================
# EXPORT ALL CONSERVATION REPORTS - EXCEL
# =========================================================

@router.get("/conservation/all/excel")
def export_all_conservation_excel(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):

    report = get_conservation_report(
        db,
        current_user
    )

    workbook = Workbook()

    default_sheet = workbook.active

    workbook.remove(
        default_sheet
    )

    summary_sheet = workbook.create_sheet(
        "Summary"
    )

    summary_sheet.append([
        "Conservation Report"
    ])

    summary_sheet.append([])

    summary_sheet.append([
        "Metric",
        "Value"
    ])

    summary = report["summary"]

    summary_sheet.append([
        "Total Conservation Statuses",
        summary["total_statuses"]
    ])

    summary_sheet.append([
        "Total Species",
        summary["total_species"]
    ])

    summary_sheet.append([
        "Total Population",
        summary["total_population"]
    ])

    summary_sheet.append([
        "Total Observations",
        summary["total_observations"]
    ])

    summary_sheet.append([
        "Total Observed Count",
        summary["total_observed_count"]
    ])

    summary_sheet.append([
        "Locations",
        summary["locations"]
    ])

    for status_report in report[
        "conservation_statuses"
    ]:

        status = status_report[
            "conservation"
        ]["status"]

        sheet_name = safe_excel_sheet_name(
            status,
            "Conservation"
        )

        sheet = workbook.create_sheet(
            sheet_name
        )

        status_summary = status_report[
            "summary"
        ]

        sheet.append([
            "Conservation Status",
            status
        ])

        sheet.append([
            "Total Species",
            status_summary[
                "total_species"
            ]
        ])

        sheet.append([
            "Total Population",
            status_summary[
                "total_population"
            ]
        ])

        sheet.append([
            "Total Observations",
            status_summary[
                "total_observations"
            ]
        ])

        sheet.append([
            "Total Observed Count",
            status_summary[
                "total_observed_count"
            ]
        ])

        sheet.append([
            "Locations",
            status_summary[
                "locations"
            ]
        ])

        sheet.append([])

        sheet.append([
            "Species",
            "Scientific Name",
            "Category",
            "Habitat",
            "Population",
            "Observations",
            "Observed Count",
            "Locations",
        ])

        for species_data in status_report[
            "species"
        ]:

            species = species_data[
                "species"
            ]

            species_summary = (
                species_data["summary"]
            )

            sheet.append([
                species["species_name"],
                species["scientific_name"],
                species["category"],
                species["habitat"],
                species["population"],
                species_summary[
                    "total_observations"
                ],
                species_summary[
                    "total_observed_count"
                ],
                species_summary[
                    "locations"
                ],
            ])

        sheet.freeze_panes = "A9"

    output = BytesIO()

    workbook.save(output)

    output.seek(0)

    return StreamingResponse(
        output,
        media_type=(
            "application/vnd.openxmlformats-"
            "officedocument.spreadsheetml.sheet"
        ),
        headers={
            "Content-Disposition":
                "attachment; filename=all-conservation-reports.xlsx"
        },
    )


# =========================================================
# EXPORT ALL CONSERVATION REPORTS - CSV
# =========================================================

@router.get("/conservation/all/csv")
def export_all_conservation_csv(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):

    report = get_conservation_report(
        db,
        current_user
    )

    output = StringIO()

    writer = csv.writer(output)

    writer.writerow([
        "Conservation Status",
        "Species",
        "Scientific Name",
        "Category",
        "Habitat",
        "Population",
        "Observations",
        "Observed Count",
        "Locations",
    ])

    for status_report in report[
        "conservation_statuses"
    ]:

        status = status_report[
            "conservation"
        ]["status"]

        for species_data in status_report[
            "species"
        ]:

            species = species_data[
                "species"
            ]

            summary = species_data[
                "summary"
            ]

            writer.writerow([
                status,

                safe_value(
                    species["species_name"]
                ),

                safe_value(
                    species["scientific_name"]
                ),

                safe_value(
                    species["category"]
                ),

                safe_value(
                    species["habitat"]
                ),

                safe_value(
                    species["population"]
                ),

                safe_value(
                    summary[
                        "total_observations"
                    ]
                ),

                safe_value(
                    summary[
                        "total_observed_count"
                    ]
                ),

                safe_value(
                    summary[
                        "locations"
                    ]
                ),
            ])

    output.seek(0)

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv; charset=utf-8",
        headers={
            "Content-Disposition":
                "attachment; filename=all-conservation-reports.csv"
        },
    )


# =========================================================
# SINGLE CONSERVATION STATUS - PDF
# =========================================================

@router.get(
    "/conservation/{conservation_status}/pdf"
)
def export_single_conservation_pdf(
    conservation_status: str,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):

    report = get_single_conservation_report(
        conservation_status,
        db,
        current_user
    )

    buffer = BytesIO()

    document = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=30,
        leftMargin=30,
        topMargin=30,
        bottomMargin=30,
    )

    styles = getSampleStyleSheet()

    story = []

    status = report[
        "conservation"
    ]["status"]

    summary = report["summary"]

    story.append(
        Paragraph(
            f"Conservation Report - "
            f"{safe_pdf_text(status)}",
            styles["Title"]
        )
    )

    story.append(
        Spacer(1, 15)
    )

    summary_data = [
        ["Metric", "Value"],

        [
            "Conservation Status",
            safe_pdf_text(status)
        ],

        [
            "Total Species",
            safe_value(
                summary["total_species"]
            )
        ],

        [
            "Total Population",
            safe_value(
                summary["total_population"]
            )
        ],

        [
            "Total Observations",
            safe_value(
                summary["total_observations"]
            )
        ],

        [
            "Total Observed Count",
            safe_value(
                summary["total_observed_count"]
            )
        ],

        [
            "Locations",
            safe_value(
                summary["locations"]
            )
        ],
    ]

    table = Table(
        summary_data,
        colWidths=[220, 200]
    )

    table.setStyle(
        TableStyle([
            (
                "BACKGROUND",
                (0, 0),
                (-1, 0),
                colors.HexColor("#dc3545")
            ),

            (
                "TEXTCOLOR",
                (0, 0),
                (-1, 0),
                colors.white
            ),

            (
                "GRID",
                (0, 0),
                (-1, -1),
                0.5,
                colors.grey
            ),

            (
                "PADDING",
                (0, 0),
                (-1, -1),
                7
            ),
        ])
    )

    story.append(table)

    story.append(
        Spacer(1, 20)
    )

    story.append(
        Paragraph(
            "Species Under This Conservation Status",
            styles["Heading2"]
        )
    )

    species_table_data = [
        [
            "Species",
            "Scientific Name",
            "Category",
            "Habitat",
            "Population",
            "Observations",
        ]
    ]

    for species_data in report[
        "species"
    ]:

        species = species_data[
            "species"
        ]

        species_summary = (
            species_data["summary"]
        )

        species_table_data.append([
            safe_pdf_text(
                species["species_name"]
            ),

            safe_pdf_text(
                species["scientific_name"]
            ),

            safe_pdf_text(
                species["category"]
            ),

            safe_pdf_text(
                species["habitat"]
            ),

            safe_value(
                species["population"]
            ),

            safe_value(
                species_summary[
                    "total_observations"
                ]
            ),
        ])

    species_table = Table(
        species_table_data,
        repeatRows=1,
    )

    species_table.setStyle(
        TableStyle([
            (
                "BACKGROUND",
                (0, 0),
                (-1, 0),
                colors.HexColor("#a71d2a")
            ),

            (
                "TEXTCOLOR",
                (0, 0),
                (-1, 0),
                colors.white
            ),

            (
                "FONTNAME",
                (0, 0),
                (-1, 0),
                "Helvetica-Bold"
            ),

            (
                "GRID",
                (0, 0),
                (-1, -1),
                0.5,
                colors.grey
            ),

            (
                "PADDING",
                (0, 0),
                (-1, -1),
                5
            ),
        ])
    )

    story.append(
        species_table
    )

    document.build(story)

    buffer.seek(0)

    filename = (
        f"{status}"
        "-conservation-report.pdf"
    )

    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={
            "Content-Disposition":
                f'attachment; filename="{filename}"'
        },
    )


# =========================================================
# SINGLE CONSERVATION STATUS - EXCEL
# =========================================================

@router.get(
    "/conservation/{conservation_status}/excel"
)
def export_single_conservation_excel(
    conservation_status: str,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):

    report = get_single_conservation_report(
        conservation_status,
        db,
        current_user
    )

    workbook = Workbook()

    sheet = workbook.active

    sheet.title = safe_excel_sheet_name(
        conservation_status
    )

    summary = report["summary"]

    sheet.append([
        "Conservation Report"
    ])

    sheet.append([])

    sheet.append([
        "Conservation Status",
        conservation_status
    ])

    sheet.append([
        "Total Species",
        summary["total_species"]
    ])

    sheet.append([
        "Total Population",
        summary["total_population"]
    ])

    sheet.append([
        "Total Observations",
        summary["total_observations"]
    ])

    sheet.append([
        "Total Observed Count",
        summary["total_observed_count"]
    ])

    sheet.append([
        "Locations",
        summary["locations"]
    ])

    sheet.append([])

    sheet.append([
        "Species",
        "Scientific Name",
        "Category",
        "Habitat",
        "Population",
        "Observations",
        "Observed Count",
        "Locations",
    ])

    for species_data in report[
        "species"
    ]:

        species = species_data[
            "species"
        ]

        species_summary = (
            species_data["summary"]
        )

        sheet.append([
            species["species_name"],
            species["scientific_name"],
            species["category"],
            species["habitat"],
            species["population"],
            species_summary[
                "total_observations"
            ],
            species_summary[
                "total_observed_count"
            ],
            species_summary[
                "locations"
            ],
        ])

    sheet.freeze_panes = "A11"

    output = BytesIO()

    workbook.save(output)

    output.seek(0)

    filename = (
        f"{conservation_status}"
        "-conservation-report.xlsx"
    )

    return StreamingResponse(
        output,
        media_type=(
            "application/vnd.openxmlformats-"
            "officedocument.spreadsheetml.sheet"
        ),
        headers={
            "Content-Disposition":
                f'attachment; filename="{filename}"'
        },
    )


# =========================================================
# SINGLE CONSERVATION STATUS - CSV
# =========================================================

@router.get(
    "/conservation/{conservation_status}/csv"
)
def export_single_conservation_csv(
    conservation_status: str,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):

    report = get_single_conservation_report(
        conservation_status,
        db,
        current_user
    )

    output = StringIO()

    writer = csv.writer(output)

    writer.writerow([
        "Conservation Status",
        "Species",
        "Scientific Name",
        "Category",
        "Habitat",
        "Population",
        "Observations",
        "Observed Count",
        "Locations",
    ])

    for species_data in report[
        "species"
    ]:

        species = species_data[
            "species"
        ]

        summary = species_data[
            "summary"
        ]

        writer.writerow([
            conservation_status,

            safe_value(
                species["species_name"]
            ),

            safe_value(
                species["scientific_name"]
            ),

            safe_value(
                species["category"]
            ),

            safe_value(
                species["habitat"]
            ),

            safe_value(
                species["population"]
            ),

            safe_value(
                summary[
                    "total_observations"
                ]
            ),

            safe_value(
                summary[
                    "total_observed_count"
                ]
            ),

            safe_value(
                summary[
                    "locations"
                ]
            ),
        ])

    output.seek(0)

    filename = (
        f"{conservation_status}"
        "-conservation-report.csv"
    )

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv; charset=utf-8",
        headers={
            "Content-Disposition":
                f'attachment; filename="{filename}"'
        },
    )