from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from io import BytesIO
import csv
import io
import re
import html

from openpyxl import Workbook

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
    PageBreak,
)

import models
from database import get_db
from auth import get_current_user


router = APIRouter(
    prefix="/reports",
    tags=["Reports"]
)


# =========================================================
# HELPER FUNCTIONS
# =========================================================

def safe_value(value, default="-"):
    """
    Convert None/empty values into a safe display value.
    """
    if value is None:
        return default

    value = str(value).strip()

    if value == "":
        return default

    return value


def safe_pdf_text(value, default="-"):
    """
    Safely convert text for ReportLab Paragraph.
    Prevents special HTML characters from breaking PDF generation.
    """
    value = safe_value(value, default)

    return html.escape(
        str(value),
        quote=True
    )


def safe_excel_sheet_name(name, fallback):
    """
    Excel sheet names:
    - Maximum 31 characters
    - Cannot contain: : \\ / ? * [ ]
    """

    if not name:
        name = fallback

    name = str(name)

    name = re.sub(
        r'[:\\/?*\[\]]',
        "_",
        name
    )

    name = name.strip()

    if not name:
        name = fallback

    return name[:31]


def build_survey_report(
    db: Session,
    survey: models.Survey
):
    """
    Build a common survey report structure.

    This structure is used by:
    - JSON report
    - Single PDF
    - Single Excel
    - Single CSV
    - All PDF
    - All Excel
    - All CSV
    """

    observations = (
        db.query(models.Observation)
        .filter(
            models.Observation.survey_id == survey.id
        )
        .order_by(
            models.Observation.observation_date.desc()
        )
        .all()
    )

    observation_data = []

    species_ids = set()
    locations = set()

    for observation in observations:

        if observation.species_id is not None:
            species_ids.add(
                observation.species_id
            )

        if observation.location:
            locations.add(
                observation.location
            )

        species_name = "Unknown Species"

        if observation.species:
            species_name = safe_value(
                observation.species.species_name,
                "Unknown Species"
            )

        observation_data.append({
            "id": observation.id,
            "species_id": observation.species_id,
            "species_name": species_name,
            "location": observation.location,
            "latitude": observation.latitude,
            "longitude": observation.longitude,
            "observation_date": observation.observation_date,
            "observer_name": observation.observer_name,
            "population_count": observation.population_count,
            "image_path": observation.image_path,
            "audio_path": observation.audio_path,
            "notes": observation.notes,
        })

    return {
        "survey": {
            "id": survey.id,
            "survey_id": survey.survey_id,
            "title": survey.title,
            "survey_date": survey.survey_date,
            "protected_area": survey.protected_area,
            "habitat_type": survey.habitat_type,
            "monitoring_location": survey.monitoring_location,
            "gps_latitude": survey.gps_latitude,
            "gps_longitude": survey.gps_longitude,
            "monitoring_device": survey.monitoring_device,
            "researcher_name": survey.researcher_name,
            "status": survey.status,
            "notes": survey.notes,
        },

        "summary": {
            "total_observations": len(observations),
            "species_recorded": len(species_ids),
            "locations": len(locations),
        },

        "observations": observation_data,
    }


# =========================================================
# GET ALL WILDLIFE SURVEY REPORT
# =========================================================

@router.get("/wildlife-survey")
def get_wildlife_survey_report(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):

    surveys = (
        db.query(models.Survey)
        .order_by(
            models.Survey.survey_date.desc()
        )
        .all()
    )

    total_observations = (
        db.query(models.Observation).count()
    )

    species_recorded = (
        db.query(models.Observation.species_id)
        .filter(
            models.Observation.species_id.isnot(None)
        )
        .distinct()
        .count()
    )

    locations = (
        db.query(models.Observation.location)
        .filter(
            models.Observation.location.isnot(None)
        )
        .distinct()
        .count()
    )

    survey_reports = []

    for survey in surveys:

        report = build_survey_report(
            db,
            survey
        )

        survey_reports.append(report)

    return {
        "report_type": "Wildlife Survey Report",

        "summary": {
            "total_surveys": len(surveys),
            "total_observations": total_observations,
            "species_recorded": species_recorded,
            "locations": locations,
        },

        "surveys": survey_reports,
    }


# =========================================================
# ALL SURVEYS - PDF
#
# IMPORTANT:
# This route MUST be before:
# /wildlife-survey/{survey_id}/pdf
#
# Frontend URL:
# /reports/wildlife-survey/all/pdf
# =========================================================

@router.get("/wildlife-survey/all/pdf")
def export_all_surveys_pdf(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):

    print("========== ALL SURVEYS PDF EXPORT ==========")

    surveys = (
        db.query(models.Survey)
        .order_by(
            models.Survey.survey_date.desc()
        )
        .all()
    )

    print(
        f"Total surveys found: {len(surveys)}"
    )

    buffer = BytesIO()

    document = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=25,
        leftMargin=25,
        topMargin=25,
        bottomMargin=25,
    )

    styles = getSampleStyleSheet()

    elements = []

    # -----------------------------------------------------
    # TITLE
    # -----------------------------------------------------

    elements.append(
        Paragraph(
            "Wildlife Survey - Complete Report",
            styles["Title"]
        )
    )

    elements.append(
        Paragraph(
            f"<b>Total Surveys:</b> "
            f"{len(surveys)}",
            styles["Normal"]
        )
    )

    elements.append(
        Spacer(1, 15)
    )

    # -----------------------------------------------------
    # OVERALL SUMMARY
    # -----------------------------------------------------

    total_observations = (
        db.query(models.Observation).count()
    )

    total_species = (
        db.query(models.Observation.species_id)
        .filter(
            models.Observation.species_id.isnot(None)
        )
        .distinct()
        .count()
    )

    total_locations = (
        db.query(models.Observation.location)
        .filter(
            models.Observation.location.isnot(None)
        )
        .distinct()
        .count()
    )

    elements.append(
        Paragraph(
            "<b>Overall Summary</b>",
            styles["Heading2"]
        )
    )

    summary_table = Table([
        ["Metric", "Value"],
        ["Total Surveys", len(surveys)],
        [
            "Total Observations",
            total_observations
        ],
        [
            "Species Recorded",
            total_species
        ],
        [
            "Monitoring Locations",
            total_locations
        ],
    ])

    summary_table.setStyle(
        TableStyle([
            (
                "BACKGROUND",
                (0, 0),
                (-1, 0),
                colors.HexColor("#198754")
            ),
            (
                "TEXTCOLOR",
                (0, 0),
                (-1, 0),
                colors.white
            ),
            (
                "GRID",
                (0, 0),
                (-1, -1),
                0.5,
                colors.grey
            ),
            (
                "BACKGROUND",
                (0, 1),
                (0, -1),
                colors.lightgrey
            ),
            (
                "PADDING",
                (0, 0),
                (-1, -1),
                6
            ),
        ])
    )

    elements.append(
        summary_table
    )

    elements.append(
        Spacer(1, 20)
    )

    # -----------------------------------------------------
    # EACH SURVEY
    # -----------------------------------------------------

    for index, survey in enumerate(surveys):

        report = build_survey_report(
            db,
            survey
        )

        elements.append(
            Paragraph(
                f"Survey {index + 1}: "
                f"{safe_pdf_text(survey.title)}",
                styles["Heading2"]
            )
        )

        elements.append(
            Paragraph(
                f"<b>Survey ID:</b> "
                f"{safe_pdf_text(survey.survey_id)}",
                styles["Normal"]
            )
        )

        elements.append(
            Spacer(1, 8)
        )

        # -------------------------------------------------
        # SURVEY DETAILS
        # -------------------------------------------------

        survey_details = [
            ["Field", "Value"],

            [
                "Survey Date",
                safe_pdf_text(
                    survey.survey_date
                )
            ],

            [
                "Protected Area",
                safe_pdf_text(
                    survey.protected_area
                )
            ],

            [
                "Habitat Type",
                safe_pdf_text(
                    survey.habitat_type
                )
            ],

            [
                "Monitoring Location",
                safe_pdf_text(
                    survey.monitoring_location
                )
            ],

            [
                "GPS",
                safe_pdf_text(
                    f"{safe_value(survey.gps_latitude)}, "
                    f"{safe_value(survey.gps_longitude)}"
                )
            ],

            [
                "Monitoring Device",
                safe_pdf_text(
                    survey.monitoring_device
                )
            ],

            [
                "Researcher",
                safe_pdf_text(
                    survey.researcher_name
                )
            ],

            [
                "Status",
                safe_pdf_text(
                    survey.status
                )
            ],

            [
                "Notes",
                safe_pdf_text(
                    survey.notes
                )
            ],
        ]

        details_table = Table(
            survey_details,
            colWidths=[150, 550]
        )

        details_table.setStyle(
            TableStyle([
                (
                    "BACKGROUND",
                    (0, 0),
                    (-1, 0),
                    colors.HexColor("#198754")
                ),
                (
                    "TEXTCOLOR",
                    (0, 0),
                    (-1, 0),
                    colors.white
                ),
                (
                    "BACKGROUND",
                    (0, 1),
                    (0, -1),
                    colors.lightgrey
                ),
                (
                    "GRID",
                    (0, 0),
                    (-1, -1),
                    0.5,
                    colors.grey
                ),
                (
                    "VALIGN",
                    (0, 0),
                    (-1, -1),
                    "TOP"
                ),
                (
                    "PADDING",
                    (0, 0),
                    (-1, -1),
                    5
                ),
                (
                    "FONTSIZE",
                    (0, 0),
                    (-1, -1),
                    8
                ),
            ])
        )

        elements.append(
            details_table
        )

        elements.append(
            Spacer(1, 12)
        )

        # -------------------------------------------------
        # SURVEY SUMMARY
        # -------------------------------------------------

        survey_summary = report["summary"]

        survey_summary_table = Table([
            [
                "Observations",
                "Species",
                "Locations"
            ],
            [
                survey_summary[
                    "total_observations"
                ],
                survey_summary[
                    "species_recorded"
                ],
                survey_summary[
                    "locations"
                ],
            ]
        ])

        survey_summary_table.setStyle(
            TableStyle([
                (
                    "BACKGROUND",
                    (0, 0),
                    (-1, 0),
                    colors.lightgrey
                ),
                (
                    "GRID",
                    (0, 0),
                    (-1, -1),
                    0.5,
                    colors.grey
                ),
                (
                    "ALIGN",
                    (0, 0),
                    (-1, -1),
                    "CENTER"
                ),
                (
                    "PADDING",
                    (0, 0),
                    (-1, -1),
                    6
                ),
            ])
        )

        elements.append(
            survey_summary_table
        )

        elements.append(
            Spacer(1, 12)
        )

        # -------------------------------------------------
        # OBSERVATIONS
        # -------------------------------------------------

        elements.append(
            Paragraph(
                "<b>Wildlife Observations</b>",
                styles["Heading3"]
            )
        )

        observation_rows = [[
            "ID",
            "Species",
            "Location",
            "Date",
            "Observer",
            "Count",
            "Latitude",
            "Longitude",
            "Notes",
        ]]

        for obs in report["observations"]:

            observation_rows.append([
                safe_value(obs["id"]),
                safe_value(
                    obs["species_name"]
                ),
                safe_value(
                    obs["location"]
                ),
                safe_value(
                    obs["observation_date"]
                ),
                safe_value(
                    obs["observer_name"]
                ),
                safe_value(
                    obs["population_count"],
                    "0"
                ),
                safe_value(
                    obs["latitude"]
                ),
                safe_value(
                    obs["longitude"]
                ),
                safe_value(
                    obs["notes"]
                ),
            ])

        if len(observation_rows) == 1:

            observation_rows.append([
                "-",
                "No observations",
                "-",
                "-",
                "-",
                "-",
                "-",
                "-",
                "-"
            ])

        observation_table = Table(
            observation_rows,
            repeatRows=1,
            colWidths=[
                30,
                100,
                100,
                65,
                90,
                45,
                60,
                60,
                150,
            ]
        )

        observation_table.setStyle(
            TableStyle([
                (
                    "BACKGROUND",
                    (0, 0),
                    (-1, 0),
                    colors.HexColor("#198754")
                ),
                (
                    "TEXTCOLOR",
                    (0, 0),
                    (-1, 0),
                    colors.white
                ),
                (
                    "GRID",
                    (0, 0),
                    (-1, -1),
                    0.4,
                    colors.grey
                ),
                (
                    "FONTSIZE",
                    (0, 0),
                    (-1, -1),
                    7
                ),
                (
                    "VALIGN",
                    (0, 0),
                    (-1, -1),
                    "TOP"
                ),
                (
                    "PADDING",
                    (0, 0),
                    (-1, -1),
                    4
                ),
            ])
        )

        elements.append(
            observation_table
        )

        if index < len(surveys) - 1:

            elements.append(
                PageBreak()
            )

    # -----------------------------------------------------
    # BUILD PDF
    # -----------------------------------------------------

    document.build(elements)

    buffer.seek(0)

    print(
        "All surveys PDF generated successfully."
    )

    print(
        "============================================"
    )

    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={
            "Content-Disposition":
                'attachment; '
                'filename="all_wildlife_surveys.pdf"'
        }
    )


# =========================================================
# ALL SURVEYS - EXCEL
#
# Frontend URL:
# /reports/wildlife-survey/all/excel
# =========================================================

@router.get("/wildlife-survey/all/excel")
def export_all_surveys_excel(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):

    print("========== ALL SURVEYS EXCEL EXPORT ==========")

    surveys = (
        db.query(models.Survey)
        .order_by(
            models.Survey.survey_date.desc()
        )
        .all()
    )

    print(
        f"Total surveys found: {len(surveys)}"
    )

    workbook = Workbook()

    # -----------------------------------------------------
    # SUMMARY SHEET
    # -----------------------------------------------------

    summary_sheet = workbook.active
    summary_sheet.title = "Summary"

    total_observations = (
        db.query(models.Observation).count()
    )

    total_species = (
        db.query(models.Observation.species_id)
        .filter(
            models.Observation.species_id.isnot(None)
        )
        .distinct()
        .count()
    )

    total_locations = (
        db.query(models.Observation.location)
        .filter(
            models.Observation.location.isnot(None)
        )
        .distinct()
        .count()
    )

    summary_sheet.append([
        "Metric",
        "Value"
    ])

    summary_sheet.append([
        "Total Surveys",
        len(surveys)
    ])

    summary_sheet.append([
        "Total Observations",
        total_observations
    ])

    summary_sheet.append([
        "Species Recorded",
        total_species
    ])

    summary_sheet.append([
        "Monitoring Locations",
        total_locations
    ])

    # -----------------------------------------------------
    # ALL SURVEYS SHEET
    # -----------------------------------------------------

    sheet = workbook.create_sheet(
        "All Surveys"
    )

    sheet.append([
        "Survey ID",
        "Survey Name",
        "Survey Date",
        "Protected Area",
        "Habitat Type",
        "Monitoring Location",
        "GPS Latitude",
        "GPS Longitude",
        "Monitoring Device",
        "Researcher",
        "Status",
        "Observation ID",
        "Species",
        "Observation Location",
        "Latitude",
        "Longitude",
        "Observation Date",
        "Observer",
        "Population Count",
        "Image Path",
        "Audio Path",
        "Notes",
    ])

    for survey in surveys:

        report = build_survey_report(
            db,
            survey
        )

        if not report["observations"]:

            sheet.append([
                survey.survey_id,
                survey.title,
                survey.survey_date,
                survey.protected_area,
                survey.habitat_type,
                survey.monitoring_location,
                survey.gps_latitude,
                survey.gps_longitude,
                survey.monitoring_device,
                survey.researcher_name,
                survey.status,
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
            ])

        else:

            for obs in report["observations"]:

                sheet.append([
                    survey.survey_id,
                    survey.title,
                    survey.survey_date,
                    survey.protected_area,
                    survey.habitat_type,
                    survey.monitoring_location,
                    survey.gps_latitude,
                    survey.gps_longitude,
                    survey.monitoring_device,
                    survey.researcher_name,
                    survey.status,

                    obs["id"],
                    obs["species_name"],
                    obs["location"],
                    obs["latitude"],
                    obs["longitude"],
                    obs["observation_date"],
                    obs["observer_name"],
                    obs["population_count"],
                    obs["image_path"],
                    obs["audio_path"],
                    obs["notes"],
                ])

    # -----------------------------------------------------
    # INDIVIDUAL SURVEY SHEETS
    # -----------------------------------------------------

    used_sheet_names = set(
        worksheet.title
        for worksheet in workbook.worksheets
    )

    for survey in surveys:

        report = build_survey_report(
            db,
            survey
        )

        base_name = (
            survey.survey_id
            or f"Survey_{survey.id}"
        )

        safe_name = safe_excel_sheet_name(
            base_name,
            f"Survey_{survey.id}"
        )

        original_name = safe_name
        counter = 1

        while safe_name in used_sheet_names:

            suffix = f"_{counter}"

            safe_name = (
                original_name[:31 - len(suffix)]
                + suffix
            )

            counter += 1

        used_sheet_names.add(
            safe_name
        )

        survey_sheet = workbook.create_sheet(
            safe_name
        )

        survey_sheet.append([
            "Field",
            "Value"
        ])

        survey_sheet.append([
            "Survey ID",
            survey.survey_id
        ])

        survey_sheet.append([
            "Survey Name",
            survey.title
        ])

        survey_sheet.append([
            "Survey Date",
            survey.survey_date
        ])

        survey_sheet.append([
            "Protected Area",
            survey.protected_area
        ])

        survey_sheet.append([
            "Habitat Type",
            survey.habitat_type
        ])

        survey_sheet.append([
            "Monitoring Location",
            survey.monitoring_location
        ])

        survey_sheet.append([
            "GPS Latitude",
            survey.gps_latitude
        ])

        survey_sheet.append([
            "GPS Longitude",
            survey.gps_longitude
        ])

        survey_sheet.append([
            "Monitoring Device",
            survey.monitoring_device
        ])

        survey_sheet.append([
            "Researcher",
            survey.researcher_name
        ])

        survey_sheet.append([
            "Status",
            survey.status
        ])

        survey_sheet.append([
            "Notes",
            survey.notes
        ])

        survey_sheet.append([])

        survey_sheet.append([
            "Total Observations",
            report["summary"][
                "total_observations"
            ]
        ])

        survey_sheet.append([
            "Species Recorded",
            report["summary"][
                "species_recorded"
            ]
        ])

        survey_sheet.append([
            "Locations",
            report["summary"][
                "locations"
            ]
        ])

        survey_sheet.append([])

        survey_sheet.append([
            "Observation ID",
            "Species",
            "Location",
            "Date",
            "Observer",
            "Population Count",
            "Latitude",
            "Longitude",
            "Notes"
        ])

        for obs in report["observations"]:

            survey_sheet.append([
                obs["id"],
                obs["species_name"],
                obs["location"],
                obs["observation_date"],
                obs["observer_name"],
                obs["population_count"],
                obs["latitude"],
                obs["longitude"],
                obs["notes"],
            ])

    # -----------------------------------------------------
    # STYLE / AUTO WIDTH
    # -----------------------------------------------------

    for worksheet in workbook.worksheets:

        for cell in worksheet[1]:

            cell.font = cell.font.copy(
                bold=True
            )

        for column in worksheet.columns:

            max_length = 0

            letter = (
                column[0].column_letter
            )

            for cell in column:

                if cell.value is not None:

                    max_length = max(
                        max_length,
                        len(str(cell.value))
                    )

            worksheet.column_dimensions[
                letter
            ].width = min(
                max_length + 2,
                45
            )

    buffer = BytesIO()

    workbook.save(buffer)

    buffer.seek(0)

    print(
        "All surveys Excel generated successfully."
    )

    print(
        "============================================"
    )

    return StreamingResponse(
        buffer,
        media_type=(
            "application/"
            "vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
        headers={
            "Content-Disposition":
                'attachment; '
                'filename="all_wildlife_surveys.xlsx"'
        }
    )


# =========================================================
# ALL SURVEYS - CSV
#
# Frontend URL:
# /reports/wildlife-survey/all/csv
# =========================================================

@router.get("/wildlife-survey/all/csv")
def export_all_surveys_csv(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):

    print("========== ALL SURVEYS CSV EXPORT ==========")

    surveys = (
        db.query(models.Survey)
        .order_by(
            models.Survey.survey_date.desc()
        )
        .all()
    )

    print(
        f"Total surveys found: {len(surveys)}"
    )

    output = io.StringIO(
        newline=""
    )

    writer = csv.writer(
        output
    )

    writer.writerow([
        "Survey ID",
        "Survey Name",
        "Survey Date",
        "Protected Area",
        "Habitat Type",
        "Monitoring Location",
        "GPS Latitude",
        "GPS Longitude",
        "Monitoring Device",
        "Researcher",
        "Status",
        "Observation ID",
        "Species",
        "Observation Location",
        "Latitude",
        "Longitude",
        "Observation Date",
        "Observer",
        "Population Count",
        "Image Path",
        "Audio Path",
        "Notes",
    ])

    for survey in surveys:

        report = build_survey_report(
            db,
            survey
        )

        if not report["observations"]:

            writer.writerow([
                survey.survey_id,
                survey.title,
                survey.survey_date,
                survey.protected_area,
                survey.habitat_type,
                survey.monitoring_location,
                survey.gps_latitude,
                survey.gps_longitude,
                survey.monitoring_device,
                survey.researcher_name,
                survey.status,
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
            ])

        else:

            for obs in report["observations"]:

                writer.writerow([
                    survey.survey_id,
                    survey.title,
                    survey.survey_date,
                    survey.protected_area,
                    survey.habitat_type,
                    survey.monitoring_location,
                    survey.gps_latitude,
                    survey.gps_longitude,
                    survey.monitoring_device,
                    survey.researcher_name,
                    survey.status,

                    obs["id"],
                    obs["species_name"],
                    obs["location"],
                    obs["latitude"],
                    obs["longitude"],
                    obs["observation_date"],
                    obs["observer_name"],
                    obs["population_count"],
                    obs["image_path"],
                    obs["audio_path"],
                    obs["notes"],
                ])

    buffer = BytesIO()

    buffer.write(
        output.getvalue().encode(
            "utf-8-sig"
        )
    )

    buffer.seek(0)

    print(
        "All surveys CSV generated successfully."
    )

    print(
        "============================================"
    )

    return StreamingResponse(
        buffer,
        media_type="text/csv",
        headers={
            "Content-Disposition":
                'attachment; '
                'filename="all_wildlife_surveys.csv"'
        }
    )


# =========================================================
# GET SINGLE SURVEY REPORT
#
# IMPORTANT:
# This comes AFTER the /all routes.
# =========================================================

@router.get("/wildlife-survey/{survey_id}")
def get_single_survey_report(
    survey_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):

    survey = (
        db.query(models.Survey)
        .filter(
            models.Survey.id == survey_id
        )
        .first()
    )

    if not survey:

        raise HTTPException(
            status_code=404,
            detail="Survey not found"
        )

    return {
        "report_type":
            "Wildlife Survey Report",

        **build_survey_report(
            db,
            survey
        )
    }


# =========================================================
# PDF - SINGLE SURVEY
# =========================================================

@router.get("/wildlife-survey/{survey_id}/pdf")
def export_single_survey_pdf(
    survey_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):

    survey = (
        db.query(models.Survey)
        .filter(
            models.Survey.id == survey_id
        )
        .first()
    )

    if not survey:

        raise HTTPException(
            status_code=404,
            detail="Survey not found"
        )

    report = build_survey_report(
        db,
        survey
    )

    buffer = BytesIO()

    document = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=30,
        leftMargin=30,
        topMargin=30,
        bottomMargin=30,
    )

    styles = getSampleStyleSheet()

    elements = []

    elements.append(
        Paragraph(
            "Wildlife Survey Report",
            styles["Title"]
        )
    )

    elements.append(
        Paragraph(
            f"<b>Survey ID:</b> "
            f"{safe_pdf_text(survey.survey_id)}",
            styles["Normal"]
        )
    )

    elements.append(
        Paragraph(
            f"<b>Survey Name:</b> "
            f"{safe_pdf_text(survey.title)}",
            styles["Normal"]
        )
    )

    elements.append(
        Spacer(1, 15)
    )

    details = [
        [
            "Survey Date",
            safe_pdf_text(
                survey.survey_date
            )
        ],
        [
            "Protected Area",
            safe_pdf_text(
                survey.protected_area
            )
        ],
        [
            "Habitat Type",
            safe_pdf_text(
                survey.habitat_type
            )
        ],
        [
            "Monitoring Location",
            safe_pdf_text(
                survey.monitoring_location
            )
        ],
        [
            "Monitoring Device",
            safe_pdf_text(
                survey.monitoring_device
            )
        ],
        [
            "Researcher",
            safe_pdf_text(
                survey.researcher_name
            )
        ],
        [
            "Status",
            safe_pdf_text(
                survey.status
            )
        ],
        [
            "GPS",
            safe_pdf_text(
                f"{safe_value(survey.gps_latitude)}, "
                f"{safe_value(survey.gps_longitude)}"
            )
        ],
        [
            "Notes",
            safe_pdf_text(
                survey.notes
            )
        ],
    ]

    table = Table(
        details,
        colWidths=[150, 350]
    )

    table.setStyle(
        TableStyle([
            (
                "BACKGROUND",
                (0, 0),
                (0, -1),
                colors.lightgrey
            ),
            (
                "GRID",
                (0, 0),
                (-1, -1),
                0.5,
                colors.grey
            ),
            (
                "VALIGN",
                (0, 0),
                (-1, -1),
                "TOP"
            ),
            (
                "PADDING",
                (0, 0),
                (-1, -1),
                6
            ),
        ])
    )

    elements.append(table)

    elements.append(
        Spacer(1, 20)
    )

    summary = report["summary"]

    elements.append(
        Paragraph(
            "<b>Survey Summary</b>",
            styles["Heading2"]
        )
    )

    summary_table = Table([
        [
            "Observations",
            summary["total_observations"]
        ],
        [
            "Species Recorded",
            summary["species_recorded"]
        ],
        [
            "Locations",
            summary["locations"]
        ],
    ])

    summary_table.setStyle(
        TableStyle([
            (
                "GRID",
                (0, 0),
                (-1, -1),
                0.5,
                colors.grey
            ),
            (
                "BACKGROUND",
                (0, 0),
                (0, -1),
                colors.lightgrey
            ),
            (
                "PADDING",
                (0, 0),
                (-1, -1),
                6
            ),
        ])
    )

    elements.append(
        summary_table
    )

    elements.append(
        Spacer(1, 20)
    )

    elements.append(
        Paragraph(
            "<b>Wildlife Observations</b>",
            styles["Heading2"]
        )
    )

    observation_rows = [[
        "ID",
        "Species",
        "Location",
        "Date",
        "Observer",
        "Count",
        "Notes",
    ]]

    for obs in report["observations"]:

        observation_rows.append([
            safe_value(obs["id"]),
            safe_value(
                obs["species_name"]
            ),
            safe_value(
                obs["location"]
            ),
            safe_value(
                obs["observation_date"]
            ),
            safe_value(
                obs["observer_name"]
            ),
            safe_value(
                obs["population_count"],
                "0"
            ),
            safe_value(
                obs["notes"]
            ),
        ])

    if len(observation_rows) == 1:

        observation_rows.append([
            "-",
            "No observations",
            "-",
            "-",
            "-",
            "-",
            "-"
        ])

    observation_table = Table(
        observation_rows,
        repeatRows=1,
        colWidths=[
            35,
            100,
            90,
            65,
            90,
            45,
            100,
        ]
    )

    observation_table.setStyle(
        TableStyle([
            (
                "BACKGROUND",
                (0, 0),
                (-1, 0),
                colors.HexColor("#198754")
            ),
            (
                "TEXTCOLOR",
                (0, 0),
                (-1, 0),
                colors.white
            ),
            (
                "GRID",
                (0, 0),
                (-1, -1),
                0.4,
                colors.grey
            ),
            (
                "FONTSIZE",
                (0, 0),
                (-1, -1),
                7
            ),
            (
                "VALIGN",
                (0, 0),
                (-1, -1),
                "TOP"
            ),
            (
                "PADDING",
                (0, 0),
                (-1, -1),
                4
            ),
        ])
    )

    elements.append(
        observation_table
    )

    document.build(elements)

    buffer.seek(0)

    filename = (
        f"{safe_value(survey.survey_id, 'survey')}"
        "_wildlife_report.pdf"
    )

    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={
            "Content-Disposition":
                f'attachment; filename="{filename}"'
        }
    )


# =========================================================
# EXCEL - SINGLE SURVEY
# =========================================================

@router.get("/wildlife-survey/{survey_id}/excel")
def export_single_survey_excel(
    survey_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):

    survey = (
        db.query(models.Survey)
        .filter(
            models.Survey.id == survey_id
        )
        .first()
    )

    if not survey:

        raise HTTPException(
            status_code=404,
            detail="Survey not found"
        )

    report = build_survey_report(
        db,
        survey
    )

    workbook = Workbook()

    # -----------------------------------------------------
    # SURVEY SHEET
    # -----------------------------------------------------

    sheet = workbook.active
    sheet.title = "Survey"

    survey_data = [
        ["Field", "Value"],

        [
            "Survey ID",
            survey.survey_id
        ],

        [
            "Survey Name",
            survey.title
        ],

        [
            "Survey Date",
            survey.survey_date
        ],

        [
            "Protected Area",
            survey.protected_area
        ],

        [
            "Habitat Type",
            survey.habitat_type
        ],

        [
            "Monitoring Location",
            survey.monitoring_location
        ],

        [
            "GPS Latitude",
            survey.gps_latitude
        ],

        [
            "GPS Longitude",
            survey.gps_longitude
        ],

        [
            "Monitoring Device",
            survey.monitoring_device
        ],

        [
            "Researcher",
            survey.researcher_name
        ],

        [
            "Status",
            survey.status
        ],

        [
            "Notes",
            survey.notes
        ],
    ]

    for row in survey_data:

        sheet.append(row)

    # -----------------------------------------------------
    # SUMMARY
    # -----------------------------------------------------

    summary_sheet = workbook.create_sheet(
        "Summary"
    )

    summary_sheet.append([
        "Metric",
        "Value"
    ])

    summary_sheet.append([
        "Total Observations",
        report["summary"][
            "total_observations"
        ]
    ])

    summary_sheet.append([
        "Species Recorded",
        report["summary"][
            "species_recorded"
        ]
    ])

    summary_sheet.append([
        "Locations",
        report["summary"][
            "locations"
        ]
    ])

    # -----------------------------------------------------
    # OBSERVATIONS
    # -----------------------------------------------------

    obs_sheet = workbook.create_sheet(
        "Observations"
    )

    obs_sheet.append([
        "ID",
        "Species ID",
        "Species",
        "Location",
        "Latitude",
        "Longitude",
        "Observation Date",
        "Observer",
        "Population Count",
        "Image Path",
        "Audio Path",
        "Notes",
    ])

    for obs in report["observations"]:

        obs_sheet.append([
            obs["id"],
            obs["species_id"],
            obs["species_name"],
            obs["location"],
            obs["latitude"],
            obs["longitude"],
            obs["observation_date"],
            obs["observer_name"],
            obs["population_count"],
            obs["image_path"],
            obs["audio_path"],
            obs["notes"],
        ])

    # -----------------------------------------------------
    # AUTO WIDTH
    # -----------------------------------------------------

    for worksheet in workbook.worksheets:

        for column in worksheet.columns:

            max_length = 0

            column_letter = (
                column[0].column_letter
            )

            for cell in column:

                if cell.value is not None:

                    max_length = max(
                        max_length,
                        len(str(cell.value))
                    )

            worksheet.column_dimensions[
                column_letter
            ].width = min(
                max_length + 2,
                45
            )

    buffer = BytesIO()

    workbook.save(buffer)

    buffer.seek(0)

    filename = (
        f"{safe_value(survey.survey_id, 'survey')}"
        "_wildlife_report.xlsx"
    )

    return StreamingResponse(
        buffer,
        media_type=(
            "application/"
            "vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
        headers={
            "Content-Disposition":
                f'attachment; filename="{filename}"'
        }
    )


# =========================================================
# CSV - SINGLE SURVEY
# =========================================================

@router.get("/wildlife-survey/{survey_id}/csv")
def export_single_survey_csv(
    survey_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):

    survey = (
        db.query(models.Survey)
        .filter(
            models.Survey.id == survey_id
        )
        .first()
    )

    if not survey:

        raise HTTPException(
            status_code=404,
            detail="Survey not found"
        )

    report = build_survey_report(
        db,
        survey
    )

    output = io.StringIO(
        newline=""
    )

    writer = csv.writer(
        output
    )

    writer.writerow([
        "Survey ID",
        "Survey Name",
        "Survey Date",
        "Protected Area",
        "Habitat Type",
        "Monitoring Location",
        "GPS Latitude",
        "GPS Longitude",
        "Researcher",
        "Status",
        "Observation ID",
        "Species",
        "Location",
        "Latitude",
        "Longitude",
        "Observation Date",
        "Observer",
        "Population Count",
        "Image Path",
        "Audio Path",
        "Notes",
    ])

    for obs in report["observations"]:

        writer.writerow([
            survey.survey_id,
            survey.title,
            survey.survey_date,
            survey.protected_area,
            survey.habitat_type,
            survey.monitoring_location,
            survey.gps_latitude,
            survey.gps_longitude,
            survey.researcher_name,
            survey.status,
            obs["id"],
            obs["species_name"],
            obs["location"],
            obs["latitude"],
            obs["longitude"],
            obs["observation_date"],
            obs["observer_name"],
            obs["population_count"],
            obs["image_path"],
            obs["audio_path"],
            obs["notes"],
        ])

    buffer = BytesIO()

    buffer.write(
        output.getvalue().encode(
            "utf-8-sig"
        )
    )

    buffer.seek(0)

    filename = (
        f"{safe_value(survey.survey_id, 'survey')}"
        "_wildlife_report.csv"
    )

    return StreamingResponse(
        buffer,
        media_type="text/csv",
        headers={
            "Content-Disposition":
                f'attachment; filename="{filename}"'
        }
    )
