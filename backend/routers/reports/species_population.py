from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from io import BytesIO
import csv
import io
import re
import html

from openpyxl import Workbook

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
    PageBreak,
)

import models
from database import get_db
from auth import get_current_user


router = APIRouter(
    prefix="/reports",
    tags=["Species Population Reports"]
)


# =========================================================
# HELPER FUNCTIONS
# =========================================================

def safe_value(value, default="-"):
    """
    Convert None/empty values into a safe display value.
    """
    if value is None:
        return default

    value = str(value).strip()

    if value == "":
        return default

    return value


def safe_pdf_text(value, default="-"):
    """
    Safely convert text for ReportLab Paragraph.
    """
    value = safe_value(value, default)

    return html.escape(
        str(value),
        quote=True
    )


def safe_excel_sheet_name(name, fallback):
    """
    Excel sheet names:
    - Maximum 31 characters
    - Cannot contain : \\ / ? * [ ]
    """

    if not name:
        name = fallback

    name = str(name)

    name = re.sub(
        r'[:\\/?*\[\]]',
        "_",
        name
    )

    name = name.strip()

    if not name:
        name = fallback

    return name[:31]


# =========================================================
# BUILD SINGLE SPECIES REPORT
# =========================================================

def build_species_population_report(
    db: Session,
    species: models.Species
):
    """
    Build common species report structure.

    Used by:
    - JSON report
    - Single PDF
    - Single Excel
    - Single CSV
    """

    observations = (
        db.query(models.Observation)
        .filter(
            models.Observation.species_id == species.id
        )
        .order_by(
            models.Observation.observation_date.desc()
        )
        .all()
    )

    observation_data = []

    total_observed_count = 0
    locations = set()

    for observation in observations:

        if observation.location:
            locations.add(
                observation.location
            )

        population_count = (
            observation.population_count
            if observation.population_count is not None
            else 0
        )

        total_observed_count += population_count

        observation_data.append({
            "id": observation.id,
            "survey_id": observation.survey_id,
            "location": observation.location,
            "latitude": observation.latitude,
            "longitude": observation.longitude,
            "observation_date": observation.observation_date,
            "observer_name": observation.observer_name,
            "population_count": observation.population_count,
            "image_path": observation.image_path,
            "audio_path": observation.audio_path,
            "notes": observation.notes,
        })

    return {
        "species": {
            "id": species.id,
            "species_name": species.species_name,
            "scientific_name": species.scientific_name,
            "category": species.category,
            "population": species.population,
            "conservation_status": species.conservation_status,
            "habitat": species.habitat,
            "image": species.image,
        },

        "summary": {
            "population": species.population,
            "total_observations": len(observations),
            "total_observed_count": total_observed_count,
            "locations": len(locations),
        },

        "observations": observation_data,
    }


# =========================================================
# GET ALL SPECIES POPULATION REPORT
# =========================================================

@router.get("/species-population")
def get_species_population_report(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):

    species_list = (
        db.query(models.Species)
        .order_by(
            models.Species.species_name.asc()
        )
        .all()
    )

    total_population = (
        db.query(
            models.Species.population
        ).all()
    )

    total_population_value = sum(
        population[0] or 0
        for population in total_population
    )

    total_observations = (
        db.query(models.Observation).count()
    )

    species_with_observations = (
        db.query(
            models.Observation.species_id
        )
        .filter(
            models.Observation.species_id.isnot(None)
        )
        .distinct()
        .count()
    )

    species_reports = []

    for species in species_list:

        report = build_species_population_report(
            db,
            species
        )

        species_reports.append(report)

    return {
        "report_type": "Species Population Report",

        "summary": {
            "total_species": len(species_list),
            "total_population": total_population_value,
            "total_observations": total_observations,
            "species_with_observations":
                species_with_observations,
        },

        "species": species_reports,
    }


# =========================================================
# GET SINGLE SPECIES REPORT
# =========================================================

@router.get("/species-population/{species_id}")
def get_single_species_report(
    species_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):

    species = (
        db.query(models.Species)
        .filter(
            models.Species.id == species_id
        )
        .first()
    )

    if not species:

        raise HTTPException(
            status_code=404,
            detail="Species not found"
        )

    return {
        "report_type":
            "Species Population Report",

        **build_species_population_report(
            db,
            species
        )
    }


# =========================================================
# ALL SPECIES - PDF
#
# Frontend URL:
# /reports/species-population/all/pdf
# =========================================================

@router.get("/species-population/all/pdf")
def export_all_species_pdf(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):

    print(
        "========== ALL SPECIES PDF EXPORT =========="
    )

    species_list = (
        db.query(models.Species)
        .order_by(
            models.Species.species_name.asc()
        )
        .all()
    )

    buffer = BytesIO()

    document = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=25,
        leftMargin=25,
        topMargin=25,
        bottomMargin=25,
    )

    styles = getSampleStyleSheet()

    elements = []

    # -----------------------------------------------------
    # TITLE
    # -----------------------------------------------------

    elements.append(
        Paragraph(
            "Species Population - Complete Report",
            styles["Title"]
        )
    )

    elements.append(
        Paragraph(
            f"<b>Total Species:</b> "
            f"{len(species_list)}",
            styles["Normal"]
        )
    )

    elements.append(
        Spacer(1, 15)
    )

    # -----------------------------------------------------
    # OVERALL SUMMARY
    # -----------------------------------------------------

    total_population = sum(
        species.population or 0
        for species in species_list
    )

    total_observations = (
        db.query(models.Observation).count()
    )

    species_with_observations = (
        db.query(
            models.Observation.species_id
        )
        .filter(
            models.Observation.species_id.isnot(None)
        )
        .distinct()
        .count()
    )

    elements.append(
        Paragraph(
            "<b>Overall Summary</b>",
            styles["Heading2"]
        )
    )

    summary_table = Table([
        ["Metric", "Value"],

        [
            "Total Species",
            len(species_list)
        ],

        [
            "Total Population",
            total_population
        ],

        [
            "Total Observations",
            total_observations
        ],

        [
            "Species With Observations",
            species_with_observations
        ],
    ])

    summary_table.setStyle(
        TableStyle([
            (
                "BACKGROUND",
                (0, 0),
                (-1, 0),
                colors.HexColor("#198754")
            ),
            (
                "TEXTCOLOR",
                (0, 0),
                (-1, 0),
                colors.white
            ),
            (
                "GRID",
                (0, 0),
                (-1, -1),
                0.5,
                colors.grey
            ),
            (
                "BACKGROUND",
                (0, 1),
                (0, -1),
                colors.lightgrey
            ),
            (
                "PADDING",
                (0, 0),
                (-1, -1),
                6
            ),
        ])
    )

    elements.append(
        summary_table
    )

    elements.append(
        Spacer(1, 20)
    )

    # -----------------------------------------------------
    # SPECIES LIST
    # -----------------------------------------------------

    elements.append(
        Paragraph(
            "<b>Species Population Overview</b>",
            styles["Heading2"]
        )
    )

    species_rows = [[
        "ID",
        "Species Name",
        "Scientific Name",
        "Category",
        "Population",
        "Conservation Status",
        "Habitat",
    ]]

    for species in species_list:

        species_rows.append([
            safe_value(species.id),
            safe_value(species.species_name),
            safe_value(species.scientific_name),
            safe_value(species.category),
            safe_value(species.population, "0"),
            safe_value(species.conservation_status),
            safe_value(species.habitat),
        ])

    if len(species_rows) == 1:

        species_rows.append([
            "-",
            "No species found",
            "-",
            "-",
            "-",
            "-",
            "-"
        ])

    species_table = Table(
        species_rows,
        repeatRows=1,
        colWidths=[
            35,
            110,
            120,
            80,
            70,
            110,
            150,
        ]
    )

    species_table.setStyle(
        TableStyle([
            (
                "BACKGROUND",
                (0, 0),
                (-1, 0),
                colors.HexColor("#198754")
            ),
            (
                "TEXTCOLOR",
                (0, 0),
                (-1, 0),
                colors.white
            ),
            (
                "GRID",
                (0, 0),
                (-1, -1),
                0.4,
                colors.grey
            ),
            (
                "FONTSIZE",
                (0, 0),
                (-1, -1),
                7
            ),
            (
                "VALIGN",
                (0, 0),
                (-1, -1),
                "TOP"
            ),
            (
                "PADDING",
                (0, 0),
                (-1, -1),
                4
            ),
        ])
    )

    elements.append(
        species_table
    )

    elements.append(
        Spacer(1, 20)
    )

    # -----------------------------------------------------
    # EACH SPECIES DETAILS
    # -----------------------------------------------------

    for index, species in enumerate(species_list):

        report = build_species_population_report(
            db,
            species
        )

        elements.append(
            Paragraph(
                f"Species {index + 1}: "
                f"{safe_pdf_text(species.species_name)}",
                styles["Heading2"]
            )
        )

        elements.append(
            Paragraph(
                f"<b>Scientific Name:</b> "
                f"{safe_pdf_text(species.scientific_name)}",
                styles["Normal"]
            )
        )

        elements.append(
            Spacer(1, 8)
        )

        # -------------------------------------------------
        # SPECIES INFORMATION
        # -------------------------------------------------

        species_details = [
            ["Field", "Value"],

            [
                "Species Name",
                safe_pdf_text(
                    species.species_name
                )
            ],

            [
                "Scientific Name",
                safe_pdf_text(
                    species.scientific_name
                )
            ],

            [
                "Category",
                safe_pdf_text(
                    species.category
                )
            ],

            [
                "Population",
                safe_pdf_text(
                    species.population,
                    "0"
                )
            ],

            [
                "Conservation Status",
                safe_pdf_text(
                    species.conservation_status
                )
            ],

            [
                "Habitat",
                safe_pdf_text(
                    species.habitat
                )
            ],

            [
                "Image",
                safe_pdf_text(
                    species.image
                )
            ],
        ]

        details_table = Table(
            species_details,
            colWidths=[160, 550]
        )

        details_table.setStyle(
            TableStyle([
                (
                    "BACKGROUND",
                    (0, 0),
                    (-1, 0),
                    colors.HexColor("#198754")
                ),
                (
                    "TEXTCOLOR",
                    (0, 0),
                    (-1, 0),
                    colors.white
                ),
                (
                    "BACKGROUND",
                    (0, 1),
                    (0, -1),
                    colors.lightgrey
                ),
                (
                    "GRID",
                    (0, 0),
                    (-1, -1),
                    0.5,
                    colors.grey
                ),
                (
                    "VALIGN",
                    (0, 0),
                    (-1, -1),
                    "TOP"
                ),
                (
                    "PADDING",
                    (0, 0),
                    (-1, -1),
                    5
                ),
                (
                    "FONTSIZE",
                    (0, 0),
                    (-1, -1),
                    8
                ),
            ])
        )

        elements.append(
            details_table
        )

        elements.append(
            Spacer(1, 12)
        )

        # -------------------------------------------------
        # SPECIES SUMMARY
        # -------------------------------------------------

        summary = report["summary"]

        species_summary_table = Table([
            [
                "Population",
                "Observations",
                "Observed Count",
                "Locations"
            ],

            [
                summary["population"],
                summary["total_observations"],
                summary["total_observed_count"],
                summary["locations"],
            ]
        ])

        species_summary_table.setStyle(
            TableStyle([
                (
                    "BACKGROUND",
                    (0, 0),
                    (-1, 0),
                    colors.lightgrey
                ),
                (
                    "GRID",
                    (0, 0),
                    (-1, -1),
                    0.5,
                    colors.grey
                ),
                (
                    "ALIGN",
                    (0, 0),
                    (-1, -1),
                    "CENTER"
                ),
                (
                    "PADDING",
                    (0, 0),
                    (-1, -1),
                    6
                ),
            ])
        )

        elements.append(
            species_summary_table
        )

        elements.append(
            Spacer(1, 12)
        )

        # -------------------------------------------------
        # OBSERVATIONS
        # -------------------------------------------------

        elements.append(
            Paragraph(
                "<b>Observation Details</b>",
                styles["Heading3"]
            )
        )

        observation_rows = [[
            "ID",
            "Location",
            "Date",
            "Observer",
            "Count",
            "Latitude",
            "Longitude",
            "Notes",
        ]]

        for obs in report["observations"]:

            observation_rows.append([
                safe_value(obs["id"]),
                safe_value(obs["location"]),
                safe_value(obs["observation_date"]),
                safe_value(obs["observer_name"]),
                safe_value(
                    obs["population_count"],
                    "0"
                ),
                safe_value(obs["latitude"]),
                safe_value(obs["longitude"]),
                safe_value(obs["notes"]),
            ])

        if len(observation_rows) == 1:

            observation_rows.append([
                "-",
                "No observations",
                "-",
                "-",
                "-",
                "-",
                "-",
                "-"
            ])

        observation_table = Table(
            observation_rows,
            repeatRows=1,
            colWidths=[
                35,
                110,
                70,
                90,
                50,
                65,
                65,
                180,
            ]
        )

        observation_table.setStyle(
            TableStyle([
                (
                    "BACKGROUND",
                    (0, 0),
                    (-1, 0),
                    colors.HexColor("#198754")
                ),
                (
                    "TEXTCOLOR",
                    (0, 0),
                    (-1, 0),
                    colors.white
                ),
                (
                    "GRID",
                    (0, 0),
                    (-1, -1),
                    0.4,
                    colors.grey
                ),
                (
                    "FONTSIZE",
                    (0, 0),
                    (-1, -1),
                    7
                ),
                (
                    "VALIGN",
                    (0, 0),
                    (-1, -1),
                    "TOP"
                ),
                (
                    "PADDING",
                    (0, 0),
                    (-1, -1),
                    4
                ),
            ])
        )

        elements.append(
            observation_table
        )

        if index < len(species_list) - 1:

            elements.append(
                PageBreak()
            )

    # -----------------------------------------------------
    # BUILD PDF
    # -----------------------------------------------------

    document.build(elements)

    buffer.seek(0)

    print(
        "All species PDF generated successfully."
    )

    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={
            "Content-Disposition":
                'attachment; '
                'filename="all_species_population_report.pdf"'
        }
    )


# =========================================================
# ALL SPECIES - EXCEL
#
# /reports/species-population/all/excel
# =========================================================

@router.get("/species-population/all/excel")
def export_all_species_excel(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):

    print(
        "========== ALL SPECIES EXCEL EXPORT =========="
    )

    species_list = (
        db.query(models.Species)
        .order_by(
            models.Species.species_name.asc()
        )
        .all()
    )

    workbook = Workbook()

    # -----------------------------------------------------
    # SUMMARY SHEET
    # -----------------------------------------------------

    summary_sheet = workbook.active
    summary_sheet.title = "Summary"

    total_population = sum(
        species.population or 0
        for species in species_list
    )

    total_observations = (
        db.query(models.Observation).count()
    )

    species_with_observations = (
        db.query(
            models.Observation.species_id
        )
        .filter(
            models.Observation.species_id.isnot(None)
        )
        .distinct()
        .count()
    )

    summary_sheet.append([
        "Metric",
        "Value"
    ])

    summary_sheet.append([
        "Total Species",
        len(species_list)
    ])

    summary_sheet.append([
        "Total Population",
        total_population
    ])

    summary_sheet.append([
        "Total Observations",
        total_observations
    ])

    summary_sheet.append([
        "Species With Observations",
        species_with_observations
    ])

    # -----------------------------------------------------
    # ALL SPECIES SHEET
    # -----------------------------------------------------

    all_species_sheet = workbook.create_sheet(
        "All Species"
    )

    all_species_sheet.append([
        "Species ID",
        "Species Name",
        "Scientific Name",
        "Category",
        "Population",
        "Conservation Status",
        "Habitat",
        "Image",
        "Total Observations",
        "Total Observed Count",
        "Locations",
    ])

    for species in species_list:

        report = build_species_population_report(
            db,
            species
        )

        summary = report["summary"]

        all_species_sheet.append([
            species.id,
            species.species_name,
            species.scientific_name,
            species.category,
            species.population,
            species.conservation_status,
            species.habitat,
            species.image,
            summary["total_observations"],
            summary["total_observed_count"],
            summary["locations"],
        ])

    # -----------------------------------------------------
    # INDIVIDUAL SPECIES SHEETS
    # -----------------------------------------------------

    used_sheet_names = set(
        worksheet.title
        for worksheet in workbook.worksheets
    )

    for species in species_list:

        report = build_species_population_report(
            db,
            species
        )

        base_name = (
            species.species_name
            or f"Species_{species.id}"
        )

        safe_name = safe_excel_sheet_name(
            base_name,
            f"Species_{species.id}"
        )

        original_name = safe_name
        counter = 1

        while safe_name in used_sheet_names:

            suffix = f"_{counter}"

            safe_name = (
                original_name[:31 - len(suffix)]
                + suffix
            )

            counter += 1

        used_sheet_names.add(
            safe_name
        )

        species_sheet = workbook.create_sheet(
            safe_name
        )

        # -------------------------------------------------
        # SPECIES INFORMATION
        # -------------------------------------------------

        species_sheet.append([
            "Field",
            "Value"
        ])

        species_sheet.append([
            "Species ID",
            species.id
        ])

        species_sheet.append([
            "Species Name",
            species.species_name
        ])

        species_sheet.append([
            "Scientific Name",
            species.scientific_name
        ])

        species_sheet.append([
            "Category",
            species.category
        ])

        species_sheet.append([
            "Population",
            species.population
        ])

        species_sheet.append([
            "Conservation Status",
            species.conservation_status
        ])

        species_sheet.append([
            "Habitat",
            species.habitat
        ])

        species_sheet.append([
            "Image",
            species.image
        ])

        species_sheet.append([])

        # -------------------------------------------------
        # SUMMARY
        # -------------------------------------------------

        species_sheet.append([
            "Total Observations",
            report["summary"][
                "total_observations"
            ]
        ])

        species_sheet.append([
            "Total Observed Count",
            report["summary"][
                "total_observed_count"
            ]
        ])

        species_sheet.append([
            "Locations",
            report["summary"][
                "locations"
            ]
        ])

        species_sheet.append([])

        # -------------------------------------------------
        # OBSERVATIONS
        # -------------------------------------------------

        species_sheet.append([
            "Observation ID",
            "Survey ID",
            "Location",
            "Latitude",
            "Longitude",
            "Observation Date",
            "Observer",
            "Population Count",
            "Image Path",
            "Audio Path",
            "Notes",
        ])

        for obs in report["observations"]:

            species_sheet.append([
                obs["id"],
                obs["survey_id"],
                obs["location"],
                obs["latitude"],
                obs["longitude"],
                obs["observation_date"],
                obs["observer_name"],
                obs["population_count"],
                obs["image_path"],
                obs["audio_path"],
                obs["notes"],
            ])

    # -----------------------------------------------------
    # STYLE / AUTO WIDTH
    # -----------------------------------------------------

    for worksheet in workbook.worksheets:

        for cell in worksheet[1]:

            cell.font = cell.font.copy(
                bold=True
            )

        for column in worksheet.columns:

            max_length = 0

            letter = (
                column[0].column_letter
            )

            for cell in column:

                if cell.value is not None:

                    max_length = max(
                        max_length,
                        len(str(cell.value))
                    )

            worksheet.column_dimensions[
                letter
            ].width = min(
                max_length + 2,
                45
            )

    buffer = BytesIO()

    workbook.save(buffer)

    buffer.seek(0)

    print(
        "All species Excel generated successfully."
    )

    return StreamingResponse(
        buffer,
        media_type=(
            "application/"
            "vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
        headers={
            "Content-Disposition":
                'attachment; '
                'filename="all_species_population_report.xlsx"'
        }
    )


# =========================================================
# ALL SPECIES - CSV
#
# /reports/species-population/all/csv
# =========================================================

@router.get("/species-population/all/csv")
def export_all_species_csv(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):

    print(
        "========== ALL SPECIES CSV EXPORT =========="
    )

    species_list = (
        db.query(models.Species)
        .order_by(
            models.Species.species_name.asc()
        )
        .all()
    )

    output = io.StringIO(
        newline=""
    )

    writer = csv.writer(
        output
    )

    writer.writerow([
        "Species ID",
        "Species Name",
        "Scientific Name",
        "Category",
        "Population",
        "Conservation Status",
        "Habitat",
        "Image",
        "Observation ID",
        "Survey ID",
        "Observation Location",
        "Latitude",
        "Longitude",
        "Observation Date",
        "Observer",
        "Population Count",
        "Image Path",
        "Audio Path",
        "Notes",
    ])

    for species in species_list:

        report = build_species_population_report(
            db,
            species
        )

        if not report["observations"]:

            writer.writerow([
                species.id,
                species.species_name,
                species.scientific_name,
                species.category,
                species.population,
                species.conservation_status,
                species.habitat,
                species.image,
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
            ])

        else:

            for obs in report["observations"]:

                writer.writerow([
                    species.id,
                    species.species_name,
                    species.scientific_name,
                    species.category,
                    species.population,
                    species.conservation_status,
                    species.habitat,
                    species.image,

                    obs["id"],
                    obs["survey_id"],
                    obs["location"],
                    obs["latitude"],
                    obs["longitude"],
                    obs["observation_date"],
                    obs["observer_name"],
                    obs["population_count"],
                    obs["image_path"],
                    obs["audio_path"],
                    obs["notes"],
                ])

    buffer = BytesIO()

    buffer.write(
        output.getvalue().encode(
            "utf-8-sig"
        )
    )

    buffer.seek(0)

    print(
        "All species CSV generated successfully."
    )

    return StreamingResponse(
        buffer,
        media_type="text/csv",
        headers={
            "Content-Disposition":
                'attachment; '
                'filename="all_species_population_report.csv"'
        }
    )


# =========================================================
# SINGLE SPECIES - PDF
#
# /reports/species-population/{species_id}/pdf
# =========================================================

@router.get("/species-population/{species_id}/pdf")
def export_single_species_pdf(
    species_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):

    species = (
        db.query(models.Species)
        .filter(
            models.Species.id == species_id
        )
        .first()
    )

    if not species:

        raise HTTPException(
            status_code=404,
            detail="Species not found"
        )

    report = build_species_population_report(
        db,
        species
    )

    buffer = BytesIO()

    document = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=30,
        leftMargin=30,
        topMargin=30,
        bottomMargin=30,
    )

    styles = getSampleStyleSheet()

    elements = []

    elements.append(
        Paragraph(
            "Species Population Report",
            styles["Title"]
        )
    )

    elements.append(
        Paragraph(
            f"<b>Species:</b> "
            f"{safe_pdf_text(species.species_name)}",
            styles["Normal"]
        )
    )

    elements.append(
        Paragraph(
            f"<b>Scientific Name:</b> "
            f"{safe_pdf_text(species.scientific_name)}",
            styles["Normal"]
        )
    )

    elements.append(
        Spacer(1, 15)
    )

    # -----------------------------------------------------
    # SPECIES INFORMATION
    # -----------------------------------------------------

    elements.append(
        Paragraph(
            "<b>Species Information</b>",
            styles["Heading2"]
        )
    )

    details = [
        [
            "Species Name",
            safe_pdf_text(
                species.species_name
            )
        ],

        [
            "Scientific Name",
            safe_pdf_text(
                species.scientific_name
            )
        ],

        [
            "Category",
            safe_pdf_text(
                species.category
            )
        ],

        [
            "Population",
            safe_pdf_text(
                species.population,
                "0"
            )
        ],

        [
            "Conservation Status",
            safe_pdf_text(
                species.conservation_status
            )
        ],

        [
            "Habitat",
            safe_pdf_text(
                species.habitat
            )
        ],

        [
            "Image",
            safe_pdf_text(
                species.image
            )
        ],
    ]

    details_table = Table(
        details,
        colWidths=[160, 350]
    )

    details_table.setStyle(
        TableStyle([
            (
                "BACKGROUND",
                (0, 0),
                (0, -1),
                colors.lightgrey
            ),
            (
                "GRID",
                (0, 0),
                (-1, -1),
                0.5,
                colors.grey
            ),
            (
                "VALIGN",
                (0, 0),
                (-1, -1),
                "TOP"
            ),
            (
                "PADDING",
                (0, 0),
                (-1, -1),
                6
            ),
        ])
    )

    elements.append(
        details_table
    )

    elements.append(
        Spacer(1, 20)
    )

    # -----------------------------------------------------
    # SUMMARY
    # -----------------------------------------------------

    summary = report["summary"]

    elements.append(
        Paragraph(
            "<b>Population Summary</b>",
            styles["Heading2"]
        )
    )

    summary_table = Table([
        [
            "Population",
            summary["population"]
        ],

        [
            "Total Observations",
            summary["total_observations"]
        ],

        [
            "Total Observed Count",
            summary["total_observed_count"]
        ],

        [
            "Locations",
            summary["locations"]
        ],
    ])

    summary_table.setStyle(
        TableStyle([
            (
                "GRID",
                (0, 0),
                (-1, -1),
                0.5,
                colors.grey
            ),
            (
                "BACKGROUND",
                (0, 0),
                (0, -1),
                colors.lightgrey
            ),
            (
                "PADDING",
                (0, 0),
                (-1, -1),
                6
            ),
        ])
    )

    elements.append(
        summary_table
    )

    elements.append(
        Spacer(1, 20)
    )

    # -----------------------------------------------------
    # OBSERVATIONS
    # -----------------------------------------------------

    elements.append(
        Paragraph(
            "<b>Observation Details</b>",
            styles["Heading2"]
        )
    )

    observation_rows = [[
        "ID",
        "Survey ID",
        "Location",
        "Date",
        "Observer",
        "Count",
        "Latitude",
        "Longitude",
        "Notes",
    ]]

    for obs in report["observations"]:

        observation_rows.append([
            safe_value(obs["id"]),
            safe_value(obs["survey_id"]),
            safe_value(obs["location"]),
            safe_value(obs["observation_date"]),
            safe_value(obs["observer_name"]),
            safe_value(
                obs["population_count"],
                "0"
            ),
            safe_value(obs["latitude"]),
            safe_value(obs["longitude"]),
            safe_value(obs["notes"]),
        ])

    if len(observation_rows) == 1:

        observation_rows.append([
            "-",
            "-",
            "No observations",
            "-",
            "-",
            "-",
            "-",
            "-",
            "-"
        ])

    observation_table = Table(
        observation_rows,
        repeatRows=1,
        colWidths=[
            30,
            55,
            80,
            55,
            75,
            40,
            55,
            55,
            100,
        ]
    )

    observation_table.setStyle(
        TableStyle([
            (
                "BACKGROUND",
                (0, 0),
                (-1, 0),
                colors.HexColor("#198754")
            ),
            (
                "TEXTCOLOR",
                (0, 0),
                (-1, 0),
                colors.white
            ),
            (
                "GRID",
                (0, 0),
                (-1, -1),
                0.4,
                colors.grey
            ),
            (
                "FONTSIZE",
                (0, 0),
                (-1, -1),
                7
            ),
            (
                "VALIGN",
                (0, 0),
                (-1, -1),
                "TOP"
            ),
            (
                "PADDING",
                (0, 0),
                (-1, -1),
                4
            ),
        ])
    )

    elements.append(
        observation_table
    )

    document.build(elements)

    buffer.seek(0)

    filename = (
        f"{safe_value(species.species_name, 'species')}"
        "_population_report.pdf"
    )

    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={
            "Content-Disposition":
                f'attachment; filename="{filename}"'
        }
    )


# =========================================================
# SINGLE SPECIES - EXCEL
#
# /reports/species-population/{species_id}/excel
# =========================================================

@router.get("/species-population/{species_id}/excel")
def export_single_species_excel(
    species_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):

    species = (
        db.query(models.Species)
        .filter(
            models.Species.id == species_id
        )
        .first()
    )

    if not species:

        raise HTTPException(
            status_code=404,
            detail="Species not found"
        )

    report = build_species_population_report(
        db,
        species
    )

    workbook = Workbook()

    # -----------------------------------------------------
    # SPECIES SHEET
    # -----------------------------------------------------

    sheet = workbook.active
    sheet.title = "Species"

    species_data = [
        ["Field", "Value"],

        [
            "Species ID",
            species.id
        ],

        [
            "Species Name",
            species.species_name
        ],

        [
            "Scientific Name",
            species.scientific_name
        ],

        [
            "Category",
            species.category
        ],

        [
            "Population",
            species.population
        ],

        [
            "Conservation Status",
            species.conservation_status
        ],

        [
            "Habitat",
            species.habitat
        ],

        [
            "Image",
            species.image
        ],
    ]

    for row in species_data:

        sheet.append(row)

    # -----------------------------------------------------
    # SUMMARY
    # -----------------------------------------------------

    summary_sheet = workbook.create_sheet(
        "Summary"
    )

    summary_sheet.append([
        "Metric",
        "Value"
    ])

    summary_sheet.append([
        "Population",
        report["summary"]["population"]
    ])

    summary_sheet.append([
        "Total Observations",
        report["summary"]["total_observations"]
    ])

    summary_sheet.append([
        "Total Observed Count",
        report["summary"]["total_observed_count"]
    ])

    summary_sheet.append([
        "Locations",
        report["summary"]["locations"]
    ])

    # -----------------------------------------------------
    # OBSERVATIONS
    # -----------------------------------------------------

    obs_sheet = workbook.create_sheet(
        "Observations"
    )

    obs_sheet.append([
        "Observation ID",
        "Survey ID",
        "Location",
        "Latitude",
        "Longitude",
        "Observation Date",
        "Observer",
        "Population Count",
        "Image Path",
        "Audio Path",
        "Notes",
    ])

    for obs in report["observations"]:

        obs_sheet.append([
            obs["id"],
            obs["survey_id"],
            obs["location"],
            obs["latitude"],
            obs["longitude"],
            obs["observation_date"],
            obs["observer_name"],
            obs["population_count"],
            obs["image_path"],
            obs["audio_path"],
            obs["notes"],
        ])

    # -----------------------------------------------------
    # AUTO WIDTH
    # -----------------------------------------------------

    for worksheet in workbook.worksheets:

        for column in worksheet.columns:

            max_length = 0

            column_letter = (
                column[0].column_letter
            )

            for cell in column:

                if cell.value is not None:

                    max_length = max(
                        max_length,
                        len(str(cell.value))
                    )

            worksheet.column_dimensions[
                column_letter
            ].width = min(
                max_length + 2,
                45
            )

    buffer = BytesIO()

    workbook.save(buffer)

    buffer.seek(0)

    filename = (
        f"{safe_value(species.species_name, 'species')}"
        "_population_report.xlsx"
    )

    return StreamingResponse(
        buffer,
        media_type=(
            "application/"
            "vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
        headers={
            "Content-Disposition":
                f'attachment; filename="{filename}"'
        }
    )


# =========================================================
# SINGLE SPECIES - CSV
#
# /reports/species-population/{species_id}/csv
# =========================================================

@router.get("/species-population/{species_id}/csv")
def export_single_species_csv(
    species_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):

    species = (
        db.query(models.Species)
        .filter(
            models.Species.id == species_id
        )
        .first()
    )

    if not species:

        raise HTTPException(
            status_code=404,
            detail="Species not found"
        )

    report = build_species_population_report(
        db,
        species
    )

    output = io.StringIO(
        newline=""
    )

    writer = csv.writer(
        output
    )

    writer.writerow([
        "Species ID",
        "Species Name",
        "Scientific Name",
        "Category",
        "Population",
        "Conservation Status",
        "Habitat",
        "Image",
        "Observation ID",
        "Survey ID",
        "Location",
        "Latitude",
        "Longitude",
        "Observation Date",
        "Observer",
        "Population Count",
        "Image Path",
        "Audio Path",
        "Notes",
    ])

    if not report["observations"]:

        writer.writerow([
            species.id,
            species.species_name,
            species.scientific_name,
            species.category,
            species.population,
            species.conservation_status,
            species.habitat,
            species.image,
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
        ])

    else:

        for obs in report["observations"]:

            writer.writerow([
                species.id,
                species.species_name,
                species.scientific_name,
                species.category,
                species.population,
                species.conservation_status,
                species.habitat,
                species.image,

                obs["id"],
                obs["survey_id"],
                obs["location"],
                obs["latitude"],
                obs["longitude"],
                obs["observation_date"],
                obs["observer_name"],
                obs["population_count"],
                obs["image_path"],
                obs["audio_path"],
                obs["notes"],
            ])

    buffer = BytesIO()

    buffer.write(
        output.getvalue().encode(
            "utf-8-sig"
        )
    )

    buffer.seek(0)

    filename = (
        f"{safe_value(species.species_name, 'species')}"
        "_population_report.csv"
    )

    return StreamingResponse(
        buffer,
        media_type="text/csv",
        headers={
            "Content-Disposition":
                f'attachment; filename="{filename}"'
        }
    )