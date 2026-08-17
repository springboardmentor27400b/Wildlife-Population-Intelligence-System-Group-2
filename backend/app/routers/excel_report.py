import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


def generate_excel_report(data: dict):

    os.makedirs("uploads/reports", exist_ok=True)

    file_path = "uploads/reports/wildlife_report.xlsx"

    workbook = Workbook()

    # =====================================================
    # SHEET 1: SUMMARY
    # =====================================================

    summary = workbook.active
    summary.title = "Summary"

    summary.append(["Wildlife Population Intelligence Report"])

    summary["A1"].font = Font(bold=True, size=16)

    summary.append([])

    summary_data = [
        ["Metric", "Value"],
        ["Population Size", data.get("population_size", 0)],
        ["Species Count", data.get("species_count", 0)],
        ["Observation Count", data.get("observation_count", 0)],
        ["Biodiversity Score", data.get("biodiversity_score", 0)],
        ["Habitat Health Score", data.get("habitat_health_score", 0)],
        ["Ecosystem Health Score", data.get("ecosystem_health_score", 0)],
        ["Population Trend", data.get("trend", "Unknown")],
    ]

    for row in summary_data:
        summary.append(row)

    # =====================================================
    # SHEET 2: SPECIES POPULATION
    # =====================================================

    species_sheet = workbook.create_sheet("Species Population")

    species_sheet.append([
        "Species",
        "Population",
        "Observations",
        "Locations"
    ])

    species_data = data.get("species_population", [])

    for item in species_data:
        species_sheet.append([
            item.get("species", "Unknown"),
            item.get("population", 0),
            item.get("observations", 0),
            item.get("locations", 0)
        ])

    # =====================================================
    # SHEET 3: BIODIVERSITY
    # =====================================================

    biodiversity_sheet = workbook.create_sheet("Biodiversity")

    biodiversity_sheet.append([
        "Species",
        "Population",
        "Diversity Contribution"
    ])

    biodiversity_data = data.get("biodiversity", [])

    for item in biodiversity_data:
        biodiversity_sheet.append([
            item.get("species", "Unknown"),
            item.get("population", 0),
            item.get("diversity_score", 0)
        ])

    # =====================================================
    # SHEET 4: HABITAT
    # =====================================================

    habitat_sheet = workbook.create_sheet("Habitat Assessment")

    habitat_sheet.append([
        "Habitat",
        "Species Count",
        "Population",
        "Habitat Health",
        "Restoration Score",
        "Priority"
    ])

    habitat_data = data.get("habitat_assessment", [])

    for item in habitat_data:
        habitat_sheet.append([
            item.get("habitat", "Unknown"),
            item.get("species_count", 0),
            item.get("population", 0),
            item.get("habitat_health_score", 0),
            item.get("restoration_score", 0),
            item.get("priority", "Unknown")
        ])

    # =====================================================
    # SHEET 5: CONSERVATION
    # =====================================================

    conservation_sheet = workbook.create_sheet("Conservation")

    conservation_sheet.append([
        "Species",
        "Population",
        "Priority Score",
        "Priority",
        "Recommendation"
    ])

    conservation_data = data.get("conservation", [])

    for item in conservation_data:
        conservation_sheet.append([
            item.get("species", "Unknown"),
            item.get("population", 0),
            item.get("priority_score", 0),
            item.get("priority", "Unknown"),
            item.get("recommendation", "")
        ])

    # =====================================================
    # FORMATTING
    # =====================================================

    for sheet in workbook.worksheets:

        # Header row
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        # Adjust column widths
        for column in sheet.columns:

            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:

                if cell.value is not None:
                    max_length = max(
                        max_length,
                        len(str(cell.value))
                    )

            sheet.column_dimensions[column_letter].width = min(
                max_length + 2,
                50
            )

    workbook.save(file_path)

    return {
        "message": "Excel report generated successfully",
        "file": file_path
    }