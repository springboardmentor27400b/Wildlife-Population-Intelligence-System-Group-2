import io
import csv
import json
import base64
import uuid
from datetime import datetime, timezone
from typing import Optional, Dict, Any, List
from fastapi.responses import StreamingResponse

import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt

from app.services.biodiversity_analytics_service import BiodiversityAnalyticsService
from app.services.report_insights_engine import ReportInsightsEngine
from app.models.user import User
from app.models.report_history import ReportHistory
from app.models.notification import Notification

class WildlifeReportService:

    @staticmethod
    def _generate_chart(title: str, labels: list, sizes: list, chart_type: str = 'pie') -> io.BytesIO:
        fig, ax = plt.subplots(figsize=(6, 4))
        if not labels:
            ax.text(0.5, 0.5, 'No Data Available', horizontalalignment='center', verticalalignment='center')
            ax.axis('off')
        else:
            if chart_type == 'pie':
                ax.pie(sizes, labels=labels, autopct='%1.1f%%', startangle=90)
                ax.axis('equal')
            elif chart_type == 'bar':
                ax.bar(labels, sizes, color='teal')
                plt.xticks(rotation=45, ha='right')
                plt.tight_layout()
            elif chart_type == 'line':
                ax.plot(labels, sizes, marker='o', color='teal', linestyle='-')
                plt.xticks(rotation=45, ha='right')
                plt.tight_layout()
        
        plt.title(title)
        buf = io.BytesIO()
        plt.savefig(buf, format='png', bbox_inches='tight')
        plt.close(fig)
        buf.seek(0)
        return buf

    @staticmethod
    def _generate_all_charts(data: Dict[str, Any]) -> Dict[str, io.BytesIO]:
        charts = {}
        
        # Species Distribution (Pie)
        sp_labels = [item['name'] for item in data['distributions']['species'][:10]]
        sp_sizes = [item['count'] for item in data['distributions']['species'][:10]]
        charts['species_distribution'] = WildlifeReportService._generate_chart('Top Species', sp_labels, sp_sizes, 'pie')
        
        # Observation Trends (Line)
        tr_labels = [item['date'][-5:] for item in data['trends']]
        tr_sizes = [item['observations'] for item in data['trends']]
        charts['observation_trends'] = WildlifeReportService._generate_chart('Observation Trends', tr_labels, tr_sizes, 'line')
        
        # Habitat Distribution (Bar)
        hb_labels = [item['name'] for item in data['distributions']['habitat']]
        hb_sizes = [item['value'] for item in data['distributions']['habitat']]
        charts['habitat_distribution'] = WildlifeReportService._generate_chart('Habitat Distribution', hb_labels, hb_sizes, 'bar')
        
        # Conservation Status (Pie)
        cs_labels = [item['name'] for item in data['distributions']['conservation_status']]
        cs_sizes = [item['value'] for item in data['distributions']['conservation_status']]
        charts['conservation_status'] = WildlifeReportService._generate_chart('Conservation Status', cs_labels, cs_sizes, 'pie')

        # Prediction Source (Pie)
        src_labels = [item['name'] for item in data['distributions']['source']]
        src_sizes = [item['value'] for item in data['distributions']['source']]
        charts['prediction_source'] = WildlifeReportService._generate_chart('Prediction Source', src_labels, src_sizes, 'pie')
        
        # Risk Analysis (Pie)
        risk_labels = [item['name'] for item in data['distributions']['risk_dashboard']]
        risk_sizes = [item['value'] for item in data['distributions']['risk_dashboard']]
        charts['risk_analysis'] = WildlifeReportService._generate_chart('Risk Analysis', risk_labels, risk_sizes, 'pie')

        return charts

    @staticmethod
    async def _log_report_history(report_type: str, user: User, filters: dict, export_format: str):
        try:
            rh = ReportHistory(
                report_type=report_type,
                user_id=str(user.id),
                user_name=user.full_name,
                filters=filters,
                export_format=export_format
            )
            await rh.insert()
            await Notification(
                title="Report Generated",
                message=f"{report_type} was successfully generated as {export_format}.",
                type="report",
                priority="Low",
                user_id=str(user.id),
                related_resource_id=str(rh.id)
            ).insert()
        except Exception:
            pass

    @staticmethod
    async def get_report_preview(
        report_type: str,
        start_date: Optional[str] = None,
        end_date: Optional[str] = None,
        species: Optional[str] = None,
        monitoring_site_id: Optional[str] = None,
        source: Optional[str] = None,
        conservation_status: Optional[str] = None,
        habitat: Optional[str] = None,
        user_name: Optional[str] = None,
        user_id: Optional[str] = None
    ) -> Dict[str, Any]:
        
        # Core Analytics
        data = await BiodiversityAnalyticsService.get_summary_analytics(
            start_date=start_date,
            end_date=end_date,
            species=species,
            source=source,
            conservation_status=conservation_status,
            habitat=habitat
        )
        
        # Generate Insights
        insights = ReportInsightsEngine.analyze(data)
        
        return {
            "metadata": {
                "report_id": str(uuid.uuid4()),
                "report_title": report_type,
                "project_name": "Wildlife Population Intelligence System",
                "milestone": "Milestone 2",
                "generated_by": user_name or "System User",
                "generated_at": datetime.now().isoformat(),
                "filters": {
                    "start_date": start_date or "All Time",
                    "end_date": end_date or "All Time",
                    "species": species or "All",
                    "source": source or "All",
                    "conservation_status": conservation_status or "All",
                    "habitat": habitat or "All"
                }
            },
            "executive_summary": {
                "text": data.get("executive_summary", "Summary unavailable."),
                "total_observations": data["summary"]["total_observations"],
                "verified_observations": data["observation_quality"]["verified"],
                "pending_observations": data["observation_quality"]["pending"],
                "total_ai_predictions": data["summary"]["total_predictions"],
                "total_species": data["summary"]["total_species"],
                "monitoring_sites": data["summary"]["active_sites"],
                "endangered_count": data["summary"]["endangered_count"],
                "average_confidence": data["summary"]["average_confidence"],
                "biodiversity_health_score": data["summary"].get("biodiversity_health_score", 0),
                "ecosystem_health_score": data["summary"].get("ecosystem_health_score", 0),
                "date_generated": datetime.now().isoformat()
            },
            "observation_statistics": {
                "trends": data["trends"],
                "quality": data.get("observation_quality", {}),
                "verified": data["observation_quality"]["verified"],
                "pending": data["observation_quality"]["pending"],
                "image_predictions": data["summary"]["image_predictions"],
                "audio_predictions": data["summary"]["audio_predictions"]
            },
            "ai_prediction_summary": {
                "image_predictions": data["summary"]["image_predictions"],
                "audio_predictions": data["summary"]["audio_predictions"],
                "source_distribution": data["distributions"]["source"],
                "performance": data.get("ai_performance", {})
            },
            "species_statistics": data["distributions"]["species"],
            "biodiversity_summary": {
                "biodiversity_score": data["summary"].get("biodiversity_health_score", 0),
                "ecosystem_health": data["summary"].get("ecosystem_health_score", 0),
                "species_richness": data["summary"]["total_species"],
                "most_detected_species": data["distributions"]["species"][0]["name"] if data["distributions"]["species"] else "None",
                "least_detected_species": data["distributions"]["species"][-1]["name"] if data["distributions"]["species"] else "None",
                "endangered_species_count": data["summary"]["endangered_count"],
                "taxonomic_categories": data["distributions"]["category"],
                "risk_dashboard": data["distributions"].get("risk_dashboard", [])
            },
            "habitat_summary": data["distributions"]["habitat"],
            "conservation_status_summary": data["distributions"]["conservation_status"],
            "alerts": data.get("alerts", []),
            "forecast_summary": data.get("forecasts", []),
            "insights": insights
        }

    @staticmethod
    async def export_excel(current_user: User, report_type: str, **filters) -> StreamingResponse:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        from openpyxl.drawing.image import Image as XLImage
        
        report_data = await WildlifeReportService.get_report_preview(report_type, **filters)
        charts = WildlifeReportService._generate_all_charts(report_data)
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Wildlife Report"
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="0d9488")
        
        ws.append([report_data["metadata"]["report_title"]])
        ws["A1"].font = Font(size=16, bold=True)
        ws.append([f"Generated At: {report_data['metadata']['generated_at'][:10]} | By: {current_user.full_name}"])
        ws.append([])
        
        # Quality & Insights
        ws.append(["Report Completeness Score:", report_data["insights"]["report_quality"]["completeness_score"]])
        ws.append([])
        
        ws.append(["Major Findings"])
        for k, v in report_data["executive_summary"].items():
            if k != "text":
                ws.append([k.replace("_", " ").title(), v])
        ws.append([])
        
        # Embed Charts
        current_row = ws.max_row + 2
        col_idx = 1
        for chart_name, chart_buf in charts.items():
            img = XLImage(chart_buf)
            ws.add_image(img, f"{openpyxl.utils.get_column_letter(col_idx)}{current_row}")
            col_idx += 8
        
        await WildlifeReportService._log_report_history(report_type, current_user, filters, "Excel")
        
        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)
        return StreamingResponse(stream, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={'Content-Disposition': f'attachment; filename="Wildlife_Report.xlsx"'})

    @staticmethod
    async def export_pdf(current_user: User, report_type: str, **filters) -> StreamingResponse:
        from fpdf import FPDF
        report_data = await WildlifeReportService.get_report_preview(report_type, **filters)
        charts = WildlifeReportService._generate_all_charts(report_data)
        
        class WildlifePDF(FPDF):
            def header(self):
                if self.page_no() > 1:
                    self.set_fill_color(13, 148, 136)
                    self.rect(0, 0, 210, 20, "F")
                    self.set_y(5)
                    self.set_font("Arial", "B", 12)
                    self.set_text_color(255, 255, 255)
                    self.cell(0, 8, report_data["metadata"]["report_title"], align="C", new_x="LMARGIN", new_y="NEXT")
                    self.ln(10)
                    self.set_text_color(0, 0, 0)
            def footer(self):
                self.set_y(-15)
                self.set_font("Arial", "I", 8)
                self.set_text_color(128, 128, 128)
                self.cell(0, 10, f"Page {self.page_no()}", align="C")

        pdf = WildlifePDF()
        
        # Cover Page
        pdf.add_page()
        pdf.set_fill_color(13, 148, 136)
        pdf.rect(0, 0, 210, 297, "F")
        pdf.set_y(100)
        pdf.set_text_color(255, 255, 255)
        pdf.set_font("Arial", "B", 28)
        pdf.cell(0, 15, "Wildlife Population", align="C", new_x="LMARGIN", new_y="NEXT")
        pdf.cell(0, 15, "Intelligence System", align="C", new_x="LMARGIN", new_y="NEXT")
        pdf.ln(10)
        pdf.set_font("Arial", "", 16)
        pdf.cell(0, 10, report_data["metadata"]["report_title"], align="C", new_x="LMARGIN", new_y="NEXT")
        pdf.ln(20)
        pdf.set_font("Arial", "", 12)
        pdf.cell(0, 8, f"Project: Default WPIS Project", align="C", new_x="LMARGIN", new_y="NEXT")
        pdf.cell(0, 8, f"Organization: WPIS Organization", align="C", new_x="LMARGIN", new_y="NEXT")
        pdf.cell(0, 8, f"Generated: {report_data['metadata']['generated_at'][:10]}", align="C", new_x="LMARGIN", new_y="NEXT")
        pdf.cell(0, 8, f"By: {current_user.full_name}", align="C", new_x="LMARGIN", new_y="NEXT")
        pdf.ln(10)
        pdf.set_font("Arial", "B", 14)
        pdf.cell(0, 8, f"Biodiversity Health Score: {report_data['executive_summary']['biodiversity_health_score']}/100", align="C", new_x="LMARGIN", new_y="NEXT")
        
        # Page 2: Exec Summary & Stats
        pdf.add_page()
        pdf.set_text_color(0, 0, 0)
        
        pdf.set_font("Arial", "B", 16)
        pdf.cell(0, 10, "Executive Summary", border="B", new_x="LMARGIN", new_y="NEXT")
        pdf.ln(2)
        pdf.set_font("Arial", "", 11)
        pdf.multi_cell(0, 6, report_data["executive_summary"]["text"])
        pdf.ln(5)
        
        pdf.set_font("Arial", "B", 14)
        pdf.cell(0, 10, "Report Statistics", new_x="LMARGIN", new_y="NEXT")
        pdf.set_font("Arial", "", 11)
        stats = report_data["insights"]["report_statistics"]
        for k, v in stats.items():
            if k != "total_pages":
                pdf.cell(90, 8, f"{k.replace('_', ' ').title()}:", border=0)
                pdf.cell(0, 8, str(v), border=0, new_x="LMARGIN", new_y="NEXT")
        pdf.ln(5)
        
        # Charts
        import tempfile
        import os
        pdf.add_page()
        pdf.set_font("Arial", "B", 16)
        pdf.cell(0, 10, "Analytical Charts", border="B", new_x="LMARGIN", new_y="NEXT")
        y_pos = pdf.get_y() + 5
        x_pos = 10
        for name, buf in charts.items():
            with tempfile.NamedTemporaryFile(delete=False, suffix='.png') as tmp:
                tmp.write(buf.getvalue())
                tmp_path = tmp.name
            if x_pos > 110:
                x_pos = 10
                y_pos += 80
            if y_pos > 220:
                pdf.add_page()
                y_pos = 20
                x_pos = 10
            pdf.image(tmp_path, x=x_pos, y=y_pos, w=85)
            x_pos += 95
            os.unlink(tmp_path)
            
        pdf.add_page()
        pdf.set_font("Arial", "B", 16)
        pdf.cell(0, 10, "Intelligent Recommendations", border="B", new_x="LMARGIN", new_y="NEXT")
        pdf.set_font("Arial", "", 11)
        for cat, recs in report_data["insights"]["recommendations"].items():
            if recs:
                pdf.set_font("Arial", "B", 12)
                pdf.cell(0, 8, cat.replace("_", " ").title(), new_x="LMARGIN", new_y="NEXT")
                pdf.set_font("Arial", "", 11)
                for rec in recs:
                    pdf.multi_cell(0, 6, f"- {rec}")
                pdf.ln(3)

        await WildlifeReportService._log_report_history(report_type, current_user, filters, "PDF")

        pdf_bytes = pdf.output(dest='S').encode('latin-1')
        stream = io.BytesIO(pdf_bytes)
        return StreamingResponse(stream, media_type="application/pdf", headers={'Content-Disposition': 'attachment; filename="Wildlife_Report.pdf"'})

    @staticmethod
    async def export_json(current_user: User, report_type: str, **filters) -> StreamingResponse:
        report_data = await WildlifeReportService.get_report_preview(report_type, **filters)
        
        charts = WildlifeReportService._generate_all_charts(report_data)
        chart_base64 = {name: base64.b64encode(buf.getvalue()).decode('utf-8') for name, buf in charts.items()}
        report_data["embedded_charts"] = chart_base64
        
        await WildlifeReportService._log_report_history(report_type, current_user, filters, "JSON")
        
        json_data = json.dumps(report_data, indent=2)
        stream = io.BytesIO(json_data.encode('utf-8'))
        return StreamingResponse(stream, media_type="application/json", headers={'Content-Disposition': 'attachment; filename="Wildlife_Report.json"'})

    @staticmethod
    async def export_csv(current_user: User, report_type: str, **filters) -> StreamingResponse:
        report_data = await WildlifeReportService.get_report_preview(report_type, **filters)
        stream = io.StringIO()
        writer = csv.writer(stream)
        
        writer.writerow(["Metadata"])
        writer.writerow(["Report Title", report_data["metadata"]["report_title"]])
        writer.writerow(["Generated At", report_data["metadata"]["generated_at"]])
        writer.writerow([])
        
        writer.writerow(["Top Species", "Count"])
        for item in report_data["species_statistics"]:
            writer.writerow([item["name"], item["count"]])
            
        await WildlifeReportService._log_report_history(report_type, current_user, filters, "CSV")
            
        bstream = io.BytesIO(stream.getvalue().encode('utf-8'))
        return StreamingResponse(bstream, media_type="text/csv", headers={'Content-Disposition': 'attachment; filename="Wildlife_Report.csv"'})

