import uuid
import os
import json
import io
import math
import hashlib
from datetime import datetime, timezone, timedelta
from typing import Optional, Dict, Any, List
from fastapi.responses import StreamingResponse
from app.database.db import supabase

class BiodiversityAnalyticsService:

    @staticmethod
    def _calculate_simpsons_index(species_counts: dict, total: int) -> float:
        if total <= 1: return 0.0
        sum_n_minus_1 = sum(count * (count - 1) for count in species_counts.values())
        d = sum_n_minus_1 / (total * (total - 1))
        return round(1 - d, 3)

    @staticmethod
    def _generate_forecasts(trends: list) -> list:
        if not trends: return []
        sorted_trends = sorted(trends, key=lambda x: x["date"])
        counts = [t["observations"] for t in sorted_trends]
        
        forecasts = []
        last_date = datetime.strptime(sorted_trends[-1]["date"], "%Y-%m-%d")
        
        window_size = min(3, len(counts))
        if window_size == 0: return []
        
        for i in range(1, 6):
            recent = counts[-window_size:]
            avg = sum(recent) / len(recent)
            counts.append(avg)
            
            next_date = (last_date + timedelta(days=i)).strftime("%Y-%m-%d")
            forecasts.append({"date": next_date, "forecast_observations": round(avg, 1)})
            
        return forecasts

    @staticmethod
    def _generate_alerts(endangered_count: int, total_observations: int) -> list:
        alerts = []
        if endangered_count > 0:
            alerts.append({"type": "critical", "message": f"{endangered_count} endangered species detections require immediate attention."})
        if total_observations > 100:
            alerts.append({"type": "info", "message": "High volume of observations recorded in the selected period."})
        if not alerts:
            alerts.append({"type": "success", "message": "Ecosystem parameters are within normal ranges."})
        return alerts

    @staticmethod
    def _generate_executive_summary(total_obs: int, total_preds: int, unique_species: int, endangered: int, avg_conf: float) -> str:
        return (f"The selected period shows {total_obs} verified field observations and {total_preds} AI predictions. "
                f"A total of {unique_species} unique species were identified with an average AI confidence of {avg_conf}%. "
                f"{'Critically, ' + str(endangered) + ' endangered species were detected, demanding conservation priority.' if endangered > 0 else 'No endangered species were detected in this period.'}")

    @staticmethod
    async def get_summary_analytics(
        start_date: Optional[str] = None,
        end_date: Optional[str] = None,
        species: Optional[str] = None,
        source: Optional[str] = None,
        conservation_status: Optional[str] = None,
        habitat: Optional[str] = None,
        category: Optional[str] = None,
        observer: Optional[str] = None,
        site_name: Optional[str] = None,
        confidence_min: Optional[float] = None,
        confidence_max: Optional[float] = None
    ) -> Dict[str, Any]:
        
        # Cache Check (5 minutes)
        query_hash = hashlib.md5(json.dumps({
            'start_date': start_date, 'end_date': end_date, 'species': species,
            'source': source, 'conservation_status': conservation_status,
            'habitat': habitat, 'category': category, 'observer': observer,
            'site_name': site_name, 'confidence_min': confidence_min, 'confidence_max': confidence_max
        }, default=str).encode()).hexdigest()
        
        try:
            res = supabase.table("advanced_analytics_cache").select("*").eq("query_hash", query_hash).execute()
            if res.data and datetime.fromisoformat(res.data[0]['expires_at']) > datetime.utcnow():
                return res.data[0]['payload']
        except Exception:
            pass

        # Build Match queries dynamically via Supabase
        def build_obs_query(q):
            if start_date: q = q.gte("observed_at", start_date)
            if end_date: q = q.lte("observed_at", end_date)
            if species: q = q.eq("species_name", species)
            if source: q = q.ilike("verification_status", f"%{source}%")
            if observer: q = q.ilike("observer_name", f"%{observer}%")
            if site_name: q = q.eq("monitoring_site_name", site_name)
            return q
            
        def build_pred_query(q):
            if start_date: q = q.gte("prediction_timestamp", start_date)
            if end_date: q = q.lte("prediction_timestamp", end_date)
            if species: q = q.eq("species_name", species)
            if source: q = q.eq("prediction_source", source)
            if conservation_status: q = q.eq("conservation_status", conservation_status)
            if habitat: q = q.eq("habitat", habitat)
            if category: q = q.eq("category", category)
            if observer: q = q.ilike("user_name", f"%{observer}%")
            if confidence_min is not None: q = q.gte("confidence_score", float(confidence_min))
            if confidence_max is not None: q = q.lte("confidence_score", float(confidence_max))
            return q

        try:
            obs_count_q = build_obs_query(supabase.table("observation_records").select("*", count="exact").limit(0))
            total_observations = obs_count_q.execute().count or 0
        except Exception:
            total_observations = 0

        try:
            pred_count_q = build_pred_query(supabase.table("unified_prediction_records").select("*", count="exact").limit(0))
            total_predictions = pred_count_q.execute().count or 0
        except Exception:
            total_predictions = 0

        try:
            sites_q = supabase.table("monitoring_sites").select("*", count="exact").eq("status", "Active").limit(0)
            active_sites = sites_q.execute().count or 0
        except Exception:
            active_sites = 0

        try:
            predictions = build_pred_query(supabase.table("unified_prediction_records").select("*")).execute().data or []
        except Exception:
            predictions = []
        
        unique_species = set()
        endangered_count = 0
        total_confidence = 0
        image_count = 0
        audio_count = 0
        
        species_counts = {}
        status_dist = {}
        habitat_dist = {}
        category_dist = {}
        source_dist = {"Image": 0, "Audio": 0}
        
        risk_dashboard = {"Low": 0, "Medium": 0, "High": 0, "Critical": 0}
        highest_conf_species = {"name": "-", "score": 0}
        lowest_conf_species = {"name": "-", "score": 100}

        for p in predictions:
            sp = p.get("species_name")
            if not sp: continue
            unique_species.add(sp)
            
            conf = p.get("confidence_score", 0)
            total_confidence += conf
            
            if sp not in species_counts:
                species_counts[sp] = {
                    'count': 0, 'total_conf': 0, 'sources': {'Image': 0, 'Audio': 0},
                    'status': p.get("conservation_status", "Unknown"),
                    'scientific_name': p.get("scientific_name", sp)
                }
            species_counts[sp]['count'] += 1
            species_counts[sp]['total_conf'] += conf
            src = p.get("prediction_source", "Unknown")
            if src in species_counts[sp]['sources']:
                species_counts[sp]['sources'][src] += 1
            
            if conf > highest_conf_species["score"]:
                highest_conf_species = {"name": sp, "score": conf}
            if conf < lowest_conf_species["score"]:
                lowest_conf_species = {"name": sp, "score": conf}
            
            if src == "Image":
                image_count += 1
                source_dist["Image"] += 1
            elif src == "Audio":
                audio_count += 1
                source_dist["Audio"] += 1
                
            status = p.get("conservation_status") or "Unknown"
            if status.lower() in ["endangered", "critically endangered"]:
                endangered_count += 1
                risk_dashboard["Critical"] += 1
            elif status.lower() in ["vulnerable", "near threatened"]:
                risk_dashboard["High"] += 1
            elif status.lower() in ["least concern"]:
                risk_dashboard["Low"] += 1
            else:
                risk_dashboard["Medium"] += 1
                
            status_dist[status] = status_dist.get(status, 0) + 1
            
            hab = p.get("habitat") or "Unknown"
            habitat_dist[hab] = habitat_dist.get(hab, 0) + 1
            
            cat = p.get("category") or "Unknown"
            category_dist[cat] = category_dist.get(cat, 0) + 1

        avg_confidence = round(total_confidence / total_predictions, 2) if total_predictions > 0 else 0
        diversity_index = BiodiversityAnalyticsService._calculate_simpsons_index({k: v['count'] for k, v in species_counts.items()}, total_predictions)
        
        biodiversity_health = min(100, max(0, int((diversity_index * 100) + (len(unique_species) * 2))))
        ecosystem_health = min(100, max(0, 100 - (risk_dashboard["Critical"] * 2)))

        top_species = sorted([{
            "name": k,
            "count": v["count"],
            "scientific_name": v["scientific_name"],
            "conservation_status": v["status"],
            "average_confidence": round(v["total_conf"] / v["count"], 2) if v["count"] > 0 else 0,
            "prediction_source": "Image" if v["sources"]["Image"] > v["sources"]["Audio"] else ("Audio" if v["sources"]["Audio"] > v["sources"]["Image"] else "Both")
        } for k, v in species_counts.items()], key=lambda x: x["count"], reverse=True)
        
        def format_dist(d):
            return [{"name": k, "value": v} for k, v in d.items()]

        try:
            observations = build_obs_query(supabase.table("observation_records").select("*")).execute().data or []
        except Exception:
            observations = []

        trend_map = {}
        verified_count = 0
        pending_count = 0
        
        for o in observations:
            v_status = o.get("verification_status", "")
            if v_status == "Verified": verified_count += 1
            elif "Pending" in v_status: pending_count += 1
                
            obs_time = o.get("observed_at")
            if obs_time:
                date_str = obs_time[:10] # YYYY-MM-DD
                trend_map[date_str] = trend_map.get(date_str, 0) + 1
                
        trends = [{"date": k, "observations": v} for k, v in sorted(trend_map.items())]
        
        forecasts = BiodiversityAnalyticsService._generate_forecasts(trends)
        alerts = BiodiversityAnalyticsService._generate_alerts(endangered_count, total_observations)
        exec_summary = BiodiversityAnalyticsService._generate_executive_summary(total_observations, total_predictions, len(unique_species), endangered_count, avg_confidence)

        payload = {
            "summary": {
                "total_observations": total_observations,
                "total_predictions": total_predictions,
                "total_species": len(unique_species),
                "endangered_count": endangered_count,
                "average_confidence": avg_confidence,
                "image_predictions": image_count,
                "audio_predictions": audio_count,
                "active_sites": active_sites,
                "biodiversity_health_score": biodiversity_health,
                "ecosystem_health_score": ecosystem_health,
                "diversity_index": diversity_index
            },
            "distributions": {
                "species": top_species,
                "conservation_status": format_dist(status_dist),
                "habitat": format_dist(habitat_dist),
                "category": format_dist(category_dist),
                "source": format_dist(source_dist),
                "risk_dashboard": format_dist(risk_dashboard)
            },
            "observation_quality": {
                "verified": verified_count,
                "pending": pending_count,
                "ai_coverage": total_predictions,
                "human_coverage": total_observations
            },
            "ai_performance": {
                "average_confidence": avg_confidence,
                "highest_confidence_species": highest_conf_species,
                "lowest_confidence_species": lowest_conf_species
            },
            "trends": trends,
            "forecasts": forecasts,
            "alerts": alerts,
            "executive_summary": exec_summary
        }
        
        try:
            supabase.table("advanced_analytics_cache").upsert({ "id": str(uuid.uuid4()),
                "query_hash": query_hash,
                "payload": payload,
                "expires_at": (datetime.utcnow() + timedelta(minutes=5)).isoformat()
            }, on_conflict="query_hash").execute()
        except Exception:
            pass

        return payload

    @staticmethod
    async def export_excel(
        start_date: Optional[str] = None,
        end_date: Optional[str] = None,
        species: Optional[str] = None,
        source: Optional[str] = None
    ) -> StreamingResponse:
        import openpyxl
        data = await BiodiversityAnalyticsService.get_summary_analytics(start_date, end_date, species, source)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Biodiversity Summary"
        ws.append(["Metric", "Value"])
        for k, v in data["summary"].items():
            ws.append([k.replace("_", " ").title(), v])
        ws.append([])
        ws.append(["Top Detected Species"])
        ws.append(["Species", "Count"])
        for item in data["distributions"]["species"]:
            ws.append([item["name"], item["count"]])
        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)
        headers = {'Content-Disposition': 'attachment; filename="biodiversity_analytics.xlsx"'}
        return StreamingResponse(stream, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)

    @staticmethod
    async def export_pdf(
        start_date: Optional[str] = None,
        end_date: Optional[str] = None,
        species: Optional[str] = None,
        source: Optional[str] = None
    ) -> StreamingResponse:
        from fpdf import FPDF
        data = await BiodiversityAnalyticsService.get_summary_analytics(start_date, end_date, species, source)
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Arial", 'B', 16)
        pdf.cell(0, 10, "Biodiversity Analytics Report", ln=True, align="C")
        pdf.set_font("Arial", 'B', 12)
        pdf.cell(0, 10, "Summary Metrics", ln=True)
        pdf.set_font("Arial", '', 10)
        for k, v in data["summary"].items():
            pdf.cell(0, 8, f"{k.replace('_', ' ').title()}: {v}", ln=True)
        pdf.cell(0, 10, "", ln=True)
        pdf.set_font("Arial", 'B', 12)
        pdf.cell(0, 10, "Executive Summary", ln=True)
        pdf.set_font("Arial", '', 10)
        pdf.multi_cell(0, 8, data["executive_summary"])
        pdf.cell(0, 10, "", ln=True)
        pdf.set_font("Arial", 'B', 12)
        pdf.cell(0, 10, "Top Detected Species", ln=True)
        pdf.set_font("Arial", '', 10)
        for item in data["distributions"]["species"][:10]:
            pdf.cell(0, 8, f"{item['name']}: {item['count']} detections", ln=True)
        pdf_bytes = pdf.output(dest='S').encode('latin-1')
        stream = io.BytesIO(pdf_bytes)
        headers = {'Content-Disposition': 'attachment; filename="biodiversity_analytics.pdf"'}
        return StreamingResponse(stream, media_type="application/pdf", headers=headers)

    @staticmethod
    async def export_json(
        start_date: Optional[str] = None,
        end_date: Optional[str] = None,
        species: Optional[str] = None,
        source: Optional[str] = None
    ) -> StreamingResponse:
        data = await BiodiversityAnalyticsService.get_summary_analytics(start_date, end_date, species, source)
        json_data = json.dumps(data, indent=2)
        stream = io.BytesIO(json_data.encode('utf-8'))
        headers = {'Content-Disposition': 'attachment; filename="biodiversity_analytics.json"'}
        return StreamingResponse(stream, media_type="application/json", headers=headers)

