import csv
import io

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer,
)
from reportlab.lib.styles import getSampleStyleSheet
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from app.models.wildlife_observation import WildlifeObservation


def get_observations(db: Session):
    return (
        db.query(WildlifeObservation)
        .order_by(WildlifeObservation.observation_date.desc())
        .all()
    )


# ============================================================
# CSV REPORT
# ============================================================

def generate_csv_report(db: Session):
    observations = get_observations(db)

    output = io.StringIO()

    writer = csv.writer(output)

    writer.writerow([
        "ID",
        "Species ID",
        "Protected Area ID",
        "Observer ID",
        "Observation Date",
        "Animal Count",
        "Latitude",
        "Longitude",
        "Observation Type",
    ])

    for obs in observations:
        writer.writerow([
            obs.id,
            obs.species_id,
            obs.protected_area_id,
            obs.observer_id,
            obs.observation_date,
            obs.animal_count,
            obs.latitude,
            obs.longitude,
            obs.observation_type,
        ])

    output.seek(0)

    return output


# ============================================================
# PDF REPORT
# ============================================================

def generate_pdf_report(db: Session):
    observations = get_observations(db)

    buffer = io.BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=25,
        leftMargin=25,
        topMargin=25,
        bottomMargin=25,
    )

    styles = getSampleStyleSheet()

    title = Paragraph(
        "Wildlife Population Intelligence System",
        styles["Title"],
    )

    subtitle = Paragraph(
        "Wildlife Observation Report",
        styles["Heading2"],
    )

    data = [[
        "ID",
        "Species",
        "Protected Area",
        "Observer",
        "Animals",
        "Date",
        "Latitude",
        "Longitude",
        "Type",
    ]]

    for obs in observations:
        data.append([
            obs.id,
            obs.species_id,
            obs.protected_area_id,
            obs.observer_id,
            obs.animal_count,
            str(obs.observation_date.date())
            if obs.observation_date
            else "",
            obs.latitude,
            obs.longitude,
            obs.observation_type,
        ])

    table = Table(
        data,
        repeatRows=1,
    )

    table.setStyle(
        TableStyle([
            (
                "BACKGROUND",
                (0, 0),
                (-1, 0),
                colors.darkgreen,
            ),
            (
                "TEXTCOLOR",
                (0, 0),
                (-1, 0),
                colors.white,
            ),
            (
                "FONTNAME",
                (0, 0),
                (-1, 0),
                "Helvetica-Bold",
            ),
            (
                "GRID",
                (0, 0),
                (-1, -1),
                0.5,
                colors.grey,
            ),
            (
                "ALIGN",
                (0, 0),
                (-1, -1),
                "CENTER",
            ),
            (
                "VALIGN",
                (0, 0),
                (-1, -1),
                "MIDDLE",
            ),
            (
                "FONTSIZE",
                (0, 0),
                (-1, -1),
                8,
            ),
        ])
    )

    elements = [
        title,
        Spacer(1, 8),
        subtitle,
        Spacer(1, 15),
        table,
    ]

    doc.build(elements)

    buffer.seek(0)

    return buffer


# ============================================================
# XLSX / EXCEL REPORT
# ============================================================

def generate_xlsx_report(db: Session):
    observations = get_observations(db)

    workbook = Workbook()

    worksheet = workbook.active
    worksheet.title = "Wildlife Observations"

    headers = [
        "ID",
        "Species ID",
        "Protected Area ID",
        "Observer ID",
        "Observation Date",
        "Animal Count",
        "Latitude",
        "Longitude",
        "Observation Type",
    ]

    worksheet.append(headers)

    # Header formatting
    header_fill = PatternFill(
        fill_type="solid",
        fgColor="1B5E20",
    )

    header_font = Font(
        bold=True,
        color="FFFFFF",
    )

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

    # Data
    for obs in observations:
        worksheet.append([
    obs.id,
    obs.species_id,
    obs.protected_area_id,
    obs.observer_id,
    obs.observation_date.replace(tzinfo=None)
    if obs.observation_date
    else None,
    obs.animal_count,
    obs.latitude,
    obs.longitude,
    obs.observation_type,
])

    # Format date column
    for cell in worksheet["E"][1:]:
        if cell.value:
            cell.number_format = "yyyy-mm-dd hh:mm:ss"

    # Make columns readable
    column_widths = {
        "A": 10,
        "B": 15,
        "C": 20,
        "D": 15,
        "E": 24,
        "F": 15,
        "G": 15,
        "H": 15,
        "I": 22,
    }

    for column, width in column_widths.items():
        worksheet.column_dimensions[column].width = width

    worksheet.freeze_panes = "A2"

    # Add filter
    worksheet.auto_filter.ref = worksheet.dimensions

    output = io.BytesIO()

    workbook.save(output)

    output.seek(0)

    return output